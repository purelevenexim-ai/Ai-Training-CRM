"""
CRM WhatsApp Routes
crm_whatsapp.py

Serves the WhatsApp messaging panel in the CRM Dashboard:
  - Template list from Wabis + Meta, normalised to a common shape
  - Send message via Wabis (primary) or Meta WA (secondary)
  - Create / submit new template via Meta WABA API

Endpoints:
  GET  /api/crm/whatsapp/templates          — fetch + normalise templates from both providers
  POST /api/crm/wabis/send                  — send via Wabis (template_id + params[])
  POST /api/crm/meta-wa/send                — send directly via Meta Graph API
  POST /api/crm/whatsapp/create-template    — submit new template to Meta WABA
"""

from __future__ import annotations

import csv
from collections import Counter
import io
import json
import logging
import os
import re
import ssl
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from typing import Any

from fastapi import APIRouter, BackgroundTasks, File, Query, UploadFile
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field

from app.config import settings
from app.services.customer_unification_service import sync_order_stats, upsert_customer
from app.services.shopify_customer_importer import ShopifyCustomerImporter
from app.storage import get_db_connection

try:
    import certifi
    _SSL_CONTEXT = ssl.create_default_context(cafile=certifi.where())
except ImportError:
    _SSL_CONTEXT = ssl.create_default_context()

logger = logging.getLogger(__name__)
router = APIRouter()

WABIS_BASE_URL = "https://bot.wabis.in/api/v1/whatsapp"
META_GRAPH_VERSION = "v20.0"
META_GRAPH_BASE = f"https://graph.facebook.com/{META_GRAPH_VERSION}"

# Simple in-process cache: {provider: (templates_list, fetched_at)}
_cache: dict[str, tuple[list[dict[str, Any]], float]] = {}
_CACHE_TTL = 1800  # 30 min
_AUDIENCE_BUCKET_PRIORITY = (
    "purchased",
    "interested",
    "abandoned",
    "whatsapp_lead",
    "promotional_lead",
)
_AUDIENCE_SUMMARY_KEY_MAP = {
    "purchased": "purchased",
    "interested": "interested",
    "abandoned": "abandoned",
    "whatsapp_lead": "whatsapp_leads",
    "promotional_lead": "promotional_leads",
}
_AUTOMATION_START_DEFAULT = "2026-05-23T00:00:00+05:30"
_IST = timezone(timedelta(hours=5, minutes=30))
_LIFECYCLE_TEMPLATE_SETTING_KEY = "whatsapp_lifecycle_template_overrides"
_LIFECYCLE_STAGE_CONFIG: dict[str, dict[str, Any]] = {
    "order_confirmed": {
        "label": "Day 1 — Order Confirmed",
        "trigger": "Starts when Shopify sends a new order-created webhook after the automation start date.",
        "description": "Operational confirmation sent immediately for a fresh order.",
        "recommended_category": "UTILITY",
        "default_templates": ["order_confirmed_v1"],
        "create_preset": {
            "name": "order_confirmed_stage_v1",
            "category": "UTILITY",
            "header": "Order Confirmed",
            "body": "Hi {{1}}, your Pureleven order {{2}} is confirmed. We will update you here when tracking is ready. ETA: {{3}}.",
            "footer": "Pureleven",
        },
    },
    "in_transit": {
        "label": "Day 2 — In Transit",
        "trigger": "Starts when Shopify fulfillment or carrier tracking is attached to the order.",
        "description": "Tracking-stage update sent once a tracking number or shipment event is available.",
        "recommended_category": "UTILITY",
        "default_templates": ["delivery_begun_v1"],
        "create_preset": {
            "name": "tracking_added_stage_v1",
            "category": "UTILITY",
            "header": "Tracking Added",
            "body": "Hi {{1}}, your Pureleven order is on the way. Tracking ID: {{2}}. ETA: {{3}}. Track here: {{4}}",
            "footer": "Pureleven",
        },
    },
    "delivered": {
        "label": "Day 5 — Delivered",
        "trigger": "Starts when the carrier webhook marks the order as delivered.",
        "description": "Delivered-stage confirmation and review invitation. VIP customers currently use a separate default template.",
        "recommended_category": "MARKETING",
        "default_templates": ["delivered_review_request_v1", "delivered_vip_v1"],
        "create_preset": {
            "name": "delivered_stage_v1",
            "category": "MARKETING",
            "header": "Delivered",
            "body": "Hi {{1}}, your Pureleven order has been delivered. If everything looks good, you can review us here: {{2}}",
            "footer": "Pureleven",
        },
    },
    "review": {
        "label": "Day 15 — Review Request",
        "trigger": "Considered during the daily 10:00 AM IST orchestration window for responsive delivered customers.",
        "description": "Review incentive or reminder stage after delivery.",
        "recommended_category": "MARKETING",
        "default_templates": ["review_incentive_v1"],
        "create_preset": {
            "name": "review_request_stage_v1",
            "category": "MARKETING",
            "header": "Share Your Review",
            "body": "Hi {{1}}, your review helps Pureleven grow. We are offering {{2}} on your next order if you review here: {{3}}. Code: {{4}}",
            "footer": "Pureleven",
        },
    },
    "upsell": {
        "label": "Day 30 — Upsell",
        "trigger": "Considered during the daily orchestration window roughly 28-33 days after delivery.",
        "description": "Cross-sell and replenishment stage using the customer’s last order context.",
        "recommended_category": "MARKETING",
        "default_templates": ["explore_products_v1"],
        "create_preset": {
            "name": "upsell_stage_v1",
            "category": "MARKETING",
            "header": "Recommended Next Order",
            "body": "Hi {{1}}, based on your last order we recommend {{2}}. Explore it here: {{3}}. Offer: {{4}}",
            "footer": "Pureleven",
        },
    },
    "corporate": {
        "label": "Day 60 — Corporate / Gifting",
        "trigger": "Considered during the daily orchestration window roughly 58-63 days after delivery.",
        "description": "Corporate or gifting follow-up. Score-based logic currently switches between high and low intent defaults.",
        "recommended_category": "MARKETING",
        "default_templates": ["corporate_pitch_high_v1", "corporate_pitch_low_v1"],
        "create_preset": {
            "name": "corporate_stage_v1",
            "category": "MARKETING",
            "header": "Corporate Gifting",
            "body": "Hi {{1}}, Pureleven can help with premium gifting and bulk orders. Your recent order value: {{2}}. Book here: {{3}}",
            "footer": "Pureleven",
        },
    },
    "loyalty": {
        "label": "Day 75 — Loyalty",
        "trigger": "Considered during the daily orchestration window roughly 73-78 days after delivery.",
        "description": "Loyalty check-in and repeat buyer reactivation for responsive customers.",
        "recommended_category": "MARKETING",
        "default_templates": ["loyalty_checkin_v1"],
        "create_preset": {
            "name": "loyalty_stage_v1",
            "category": "MARKETING",
            "header": "Loyalty Check-in",
            "body": "Hi {{1}}, it has been {{2}} days since your last order. Come back here: {{3}} and use {{4}}",
            "footer": "Pureleven",
        },
    },
    "rfm": {
        "label": "Day 95 — Re-engagement",
        "trigger": "Considered during the daily orchestration window roughly 93-98 days after delivery.",
        "description": "RFM-based winback. Current defaults vary by VIP, responsive, low, or dormant segment.",
        "recommended_category": "MARKETING",
        "default_templates": ["vip_exclusive_v1", "repeat_buyer_exclusive_v1", "winback_offer_v1", "extreme_winback_v1"],
        "create_preset": {
            "name": "reengagement_stage_v1",
            "category": "MARKETING",
            "header": "Come Back to Pureleven",
            "body": "Hi {{1}}, we miss you. Revisit Pureleven here: {{2}} and use {{3}}. Special offer: {{4}}",
            "footer": "Pureleven",
        },
    },
}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _http_get(url: str, headers: dict[str, str] | None = None, timeout: int = 10) -> dict[str, Any]:
    req = urllib.request.Request(url, headers={
        "Accept": "application/json",
        "User-Agent": "PurelevenCRM/1.0",
        **(headers or {}),
    })
    with urllib.request.urlopen(req, timeout=timeout, context=_SSL_CONTEXT) as resp:  # noqa: S310
        return json.loads(resp.read())


def _http_post(url: str, payload: dict[str, Any], headers: dict[str, str] | None = None, timeout: int = 15) -> dict[str, Any]:
    data = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=data, headers={
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "PurelevenCRM/1.0",
        **(headers or {}),
    }, method="POST")
    with urllib.request.urlopen(req, timeout=timeout, context=_SSL_CONTEXT) as resp:  # noqa: S310
        return json.loads(resp.read())


def _count_vars(text: str) -> int:
    """Count number of unique {{N}} placeholders in text."""
    nums = set(re.findall(r'\{\{(\d+)\}\}', text or ''))
    return len(nums)


def _parse_component_vars(comp: dict[str, Any]) -> tuple[int, list[str]]:
    """Return (var_count, examples) for a template component dict (Meta format)."""
    text = comp.get("text", "")
    url = comp.get("url", "")
    n = max(_count_vars(text), _count_vars(url))
    examples: list[str] = []
    ep = comp.get("example", {})
    if isinstance(ep, dict):
        body_text = ep.get("body_text", [[]])[0] if "body_text" in ep else []
        header_text = ep.get("header_text", [])
        examples = body_text or header_text or []
    elif isinstance(ep, list):
        examples = [str(value) for value in ep]
    return n, examples


def _default_meta_examples(count: int, prefix: str = "Sample") -> list[str]:
    return [f"{prefix} {index}" for index in range(1, count + 1)]


def _normalise_wabis_template(raw: dict[str, Any]) -> dict[str, Any]:
    """
    Convert a raw Wabis template record into the normalised shape expected by
    the WhatsAppPanel frontend.

    Wabis stores the Meta component definition inside `template_json` as a string.
    """
    tid = raw.get("id", 0)
    name = raw.get("template_name", "")
    category = raw.get("category", "MARKETING").upper()
    locale = raw.get("language_code") or raw.get("language") or "en_US"
    status_str = (raw.get("status") or "").upper()

    # Parse embedded template_json
    tpl_json_raw = raw.get("template_json") or raw.get("components_json") or ""
    if isinstance(tpl_json_raw, str):
        try:
            tpl_json = json.loads(tpl_json_raw)
        except Exception:
            tpl_json = {}
    else:
        tpl_json = tpl_json_raw if isinstance(tpl_json_raw, dict) else {}

    components: list[dict[str, Any]] = tpl_json.get("components", []) if tpl_json else []

    # If template_json has no components, look at top-level keys
    if not components and raw.get("components"):
        raw_comps = raw["components"]
        if isinstance(raw_comps, str):
            try:
                components = json.loads(raw_comps)
            except Exception:
                components = []
        elif isinstance(raw_comps, list):
            components = raw_comps

    header: dict[str, Any] = {}
    body: dict[str, Any] = {}
    footer: dict[str, Any] = {}
    buttons: list[dict[str, Any]] = []
    total_vars = 0
    var_labels: dict[str, str] = {}

    for comp in components:
        comp_type = (comp.get("type") or "").upper()
        if comp_type == "HEADER":
            fmt = (comp.get("format") or "TEXT").upper()
            h_text = comp.get("text", "")
            h_vars, h_ex = _parse_component_vars(comp)
            total_vars += h_vars
            header = {"text": h_text, "format": fmt, "vars": h_vars, "examples": h_ex}
        elif comp_type == "BODY":
            b_text = comp.get("text", "")
            b_vars, b_ex = _parse_component_vars(comp)
            total_vars += b_vars
            body = {"meta_text": b_text, "text": b_text, "vars": b_vars, "examples": b_ex}
        elif comp_type == "FOOTER":
            footer = {"text": comp.get("text", "")}
        elif comp_type == "BUTTONS":
            btn_list = comp.get("buttons", [])
            for bi, btn in enumerate(btn_list):
                sub = (btn.get("type") or "").upper()
                bvars, bex = _parse_component_vars(btn)
                total_vars += bvars
                buttons.append({
                    "text": btn.get("text", ""),
                    "sub_type": sub,
                    "vars": bvars,
                    "examples": bex,
                    "url": btn.get("url", ""),
                    "phone_number": btn.get("phone_number", ""),
                })

    # Fallback: if components parsing failed, try raw body_text field
    if not body and raw.get("body"):
        b_text = raw["body"]
        b_vars = _count_vars(b_text)
        body = {"meta_text": b_text, "text": b_text, "vars": b_vars, "examples": []}
        total_vars = b_vars

    # send_issue detection
    send_issue = None
    hfmt = (header.get("format") or "TEXT").upper()
    if hfmt in ("IMAGE", "VIDEO", "DOCUMENT"):
        send_issue = "image_header"

    # Locale mismatch: check if stored locale conflicts with template name
    known_locales = {"en_us": "en_US", "en_gb": "en_GB", "en_in": "en_IN"}
    loc_norm = known_locales.get(locale.lower().replace("-", "_"), locale)
    if loc_norm.lower() not in ("en_us", "en_gb", "en_in", "en") and send_issue is None:
        send_issue = "locale_mismatch"

    return {
        "id": tid,
        "name": name,
        "category": category,
        "locale": loc_norm,
        "status": status_str or "APPROVED",
        "total_vars": total_vars,
        "var_labels": var_labels,
        "header": header if header else None,
        "body": body if body else None,
        "footer": footer if footer else None,
        "buttons": buttons,
        "header_format": hfmt.lower() if header else "none",
        "send_issue": send_issue,
    }


def _normalise_meta_template(raw: dict[str, Any]) -> dict[str, Any]:
    """Convert a raw Meta WABA template list item into the normalised shape."""
    name = raw.get("name", "")
    category = (raw.get("category") or "MARKETING").upper()
    locale = (raw.get("language") or "en_US")
    status_str = (raw.get("status") or "").upper()

    components: list[dict[str, Any]] = raw.get("components", [])

    header: dict[str, Any] = {}
    body: dict[str, Any] = {}
    footer: dict[str, Any] = {}
    buttons: list[dict[str, Any]] = []
    total_vars = 0

    for comp in components:
        comp_type = (comp.get("type") or "").upper()
        if comp_type == "HEADER":
            fmt = (comp.get("format") or "TEXT").upper()
            h_text = comp.get("text", "")
            h_vars, h_ex = _parse_component_vars(comp)
            total_vars += h_vars
            header = {"text": h_text, "format": fmt, "vars": h_vars, "examples": h_ex}
        elif comp_type == "BODY":
            b_text = comp.get("text", "")
            b_vars, b_ex = _parse_component_vars(comp)
            total_vars += b_vars
            body = {"meta_text": b_text, "text": b_text, "vars": b_vars, "examples": b_ex}
        elif comp_type == "FOOTER":
            footer = {"text": comp.get("text", "")}
        elif comp_type == "BUTTONS":
            btn_list = comp.get("buttons", [])
            for btn in btn_list:
                sub = (btn.get("type") or "").upper()
                bvars, bex = _parse_component_vars(btn)
                total_vars += bvars
                buttons.append({
                    "text": btn.get("text", ""),
                    "sub_type": sub,
                    "vars": bvars,
                    "examples": bex,
                    "url": btn.get("url", ""),
                    "phone_number": btn.get("phone_number", ""),
                })

    send_issue = None
    hfmt = (header.get("format") or "TEXT").upper()
    if hfmt in ("IMAGE", "VIDEO", "DOCUMENT"):
        send_issue = "image_header"
    if status_str not in ("APPROVED", ""):
        send_issue = send_issue or "locale_mismatch"

    return {
        "id": raw.get("id", name),
        "name": name,
        "category": category,
        "locale": locale,
        "status": status_str,
        "total_vars": total_vars,
        "var_labels": {},
        "header": header if header else None,
        "body": body if body else None,
        "footer": footer if footer else None,
        "buttons": buttons,
        "header_format": hfmt.lower() if header else "none",
        "send_issue": send_issue,
    }


# ─── Template Fetch ───────────────────────────────────────────────────────────

def _fetch_wabis_templates() -> list[dict[str, Any]]:
    """Fetch and normalise all templates from Wabis."""
    api_key = settings.wabis_api_key
    if not api_key:
        logger.warning("WABIS_API_KEY not set — returning empty template list")
        return []

    phone_number_id = settings.wabis_phone_number_id
    if not phone_number_id:
        logger.warning("WABIS_PHONE_NUMBER_ID not set — returning empty template list")
        return []

    params = urllib.parse.urlencode({
        "apiToken": api_key,
        "phone_number_id": phone_number_id,
    })
    url = f"{WABIS_BASE_URL}/template/list?{params}"
    try:
        data = _http_get(url)
        raw_list = data.get("message", [])
        if not isinstance(raw_list, list):
            raw_list = []
        return [_normalise_wabis_template(t) for t in raw_list]
    except Exception as exc:
        logger.warning("Wabis template fetch error: %s", exc)
        return []


def _fetch_meta_templates() -> list[dict[str, Any]]:
    """Fetch and normalise all templates from Meta WABA."""
    token = settings.meta_access_token
    waba_id = getattr(settings, "meta_waba_id", None) or ""
    if not token or not waba_id:
        return []
    url = f"{META_GRAPH_BASE}/{waba_id}/message_templates?limit=200&fields=name,category,language,status,components"
    try:
        data = _http_get(url, headers={"Authorization": f"Bearer {token}"})
        raw_list = data.get("data", [])
        return [_normalise_meta_template(t) for t in raw_list if isinstance(t, dict)]
    except Exception as exc:
        logger.warning("Meta template fetch error: %s", exc)
        return []


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _automation_start_dt() -> datetime:
    raw = (os.getenv("WHATSAPP_AUTOMATION_START_ISO") or _AUTOMATION_START_DEFAULT).strip()
    parsed = _parse_datetime_like(raw)
    if parsed is None:
        parsed = datetime(2026, 5, 23, 0, 0, tzinfo=_IST)
    return parsed.astimezone(timezone.utc)


def _automation_start_iso() -> str:
    return _automation_start_dt().isoformat()


def _automation_start_label() -> str:
    return _automation_start_dt().astimezone(_IST).strftime("%d %b %Y, %I:%M %p IST")


def _json_loads(value: str | None, default: Any) -> Any:
    if not value:
        return default
    try:
        return json.loads(value)
    except Exception:
        return default


def _normalise_phone(phone: str | None) -> str:
    digits = re.sub(r"\D", "", phone or "")
    if not digits:
        return ""
    if len(digits) == 10:
        return f"+91{digits}"
    if len(digits) == 12 and digits.startswith("91"):
        return f"+{digits}"
    if digits.startswith("0") and len(digits) > 10:
        return f"+{digits.lstrip('0')}"
    return f"+{digits}"


def _display_email(email: str | None) -> str | None:
    if not email:
        return None
    return None if email.endswith("@whatsapp.local") else email


def _journey_scope(alias: str = "jc") -> tuple[str, list[Any]]:
    shop_domain = (getattr(settings, "default_shop_domain", "") or "").strip()
    if not shop_domain:
        return "1=1", []
    return f"{alias}.shop_domain = ?", [shop_domain]


def _get_cached_templates(provider: str) -> list[dict[str, Any]]:
    now = time.time()
    cached = _cache.get(provider)
    if cached and (now - cached[1]) < _CACHE_TTL:
        return cached[0]

    if provider == "wabis":
        fresh = _fetch_wabis_templates()
    else:
        fresh = _fetch_meta_templates()
    _cache[provider] = (fresh, now)
    return fresh


def _load_json_setting(conn: Any, setting_key: str) -> dict[str, Any]:
    row = conn.execute(
        "SELECT setting_value FROM app_settings WHERE setting_key = ?",
        (setting_key,),
    ).fetchone()
    if not row or not row["setting_value"]:
        return {}
    try:
        payload = json.loads(row["setting_value"])
    except json.JSONDecodeError:
        return {}
    return payload if isinstance(payload, dict) else {}


def _save_json_setting(conn: Any, setting_key: str, payload: dict[str, Any]) -> None:
    updated_at = datetime.now(timezone.utc).isoformat()
    conn.execute(
        """
        INSERT INTO app_settings (setting_key, setting_value, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(setting_key) DO UPDATE SET
            setting_value = excluded.setting_value,
            updated_at = excluded.updated_at
        """,
        (setting_key, json.dumps(payload, ensure_ascii=True, sort_keys=True), updated_at),
    )


def _combined_template_catalog(force: bool = False) -> dict[str, dict[str, Any]]:
    if force:
        now = time.time()
        wabis_templates = _fetch_wabis_templates()
        meta_templates = _fetch_meta_templates()
        _cache["wabis"] = (wabis_templates, now)
        _cache["meta"] = (meta_templates, now)
    else:
        wabis_templates = _get_cached_templates("wabis")
        meta_templates = _get_cached_templates("meta")

    catalog: dict[str, dict[str, Any]] = {}
    for template in wabis_templates:
        name = str(template.get("name") or "").strip()
        if not name:
            continue
        catalog[name] = {**template, "_provider": "wabis"}
    for template in meta_templates:
        name = str(template.get("name") or "").strip()
        if not name:
            continue
        catalog[name] = {**template, "_provider": "meta"}
    return catalog


def _template_signature(template: dict[str, Any]) -> dict[str, Any]:
    return {
        "header_vars": int((template.get("header") or {}).get("vars") or 0),
        "body_vars": int((template.get("body") or {}).get("vars") or 0),
        "buttons": [
            {
                "index": index,
                "sub_type": str(button.get("sub_type") or button.get("type") or "url").lower(),
                "vars": int(button.get("vars") or 0),
            }
            for index, button in enumerate(template.get("buttons") or [])
            if int(button.get("vars") or 0) > 0
        ],
    }


def _lifecycle_stage_payload(stage_key: str, overrides: dict[str, Any], catalog: dict[str, dict[str, Any]]) -> dict[str, Any]:
    config = _LIFECYCLE_STAGE_CONFIG[stage_key]
    override = overrides.get(stage_key) if isinstance(overrides, dict) else None
    active_template_name = str((override or {}).get("template_name") or config["default_templates"][0])
    return {
        "stage": stage_key,
        "label": config["label"],
        "trigger": config["trigger"],
        "description": config["description"],
        "recommended_category": config["recommended_category"],
        "default_template_names": config["default_templates"],
        "default_templates_available": [name for name in config["default_templates"] if name in catalog],
        "uses_segment_defaults": len(config["default_templates"]) > 1,
        "active_template_name": active_template_name,
        "active_template_found": active_template_name in catalog,
        "override": override or None,
        "create_preset": config["create_preset"],
    }


class _LifecycleTemplateStageRequest(BaseModel):
    stage: str
    template_name: str = ""


@router.get("/crm/whatsapp/lifecycle-templates")
def get_lifecycle_stage_templates(force: bool = False) -> JSONResponse:
    catalog = _combined_template_catalog(force=force)
    with get_db_connection() as conn:
        overrides = _load_json_setting(conn, _LIFECYCLE_TEMPLATE_SETTING_KEY)

    stages = [
        _lifecycle_stage_payload(stage_key, overrides, catalog)
        for stage_key in _LIFECYCLE_STAGE_CONFIG
    ]
    return JSONResponse({
        "ok": True,
        "stages": stages,
        "templates_loaded": len(catalog),
    })


@router.post("/crm/whatsapp/lifecycle-templates")
def save_lifecycle_stage_template(req: _LifecycleTemplateStageRequest) -> JSONResponse:
    stage_key = str(req.stage or "").strip().lower()
    if stage_key not in _LIFECYCLE_STAGE_CONFIG:
        return JSONResponse({"ok": False, "detail": "Unsupported lifecycle stage"}, status_code=400)

    catalog = _combined_template_catalog(force=False)
    template_name = str(req.template_name or "").strip()

    with get_db_connection() as conn:
        overrides = _load_json_setting(conn, _LIFECYCLE_TEMPLATE_SETTING_KEY)

        if not template_name:
            overrides.pop(stage_key, None)
        else:
            template = catalog.get(template_name)
            if not template:
                return JSONResponse({"ok": False, "detail": "Template not found in the synced library"}, status_code=404)

            signature = _template_signature(template)
            if int(signature.get("header_vars") or 0) > 0:
                return JSONResponse({"ok": False, "detail": "Templates with dynamic header variables are not supported for lifecycle stage overrides yet"}, status_code=400)

            overrides[stage_key] = {
                "template_name": template_name,
                "provider": template.get("_provider") or "meta",
                "signature": signature,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }

        _save_json_setting(conn, _LIFECYCLE_TEMPLATE_SETTING_KEY, overrides)

    return JSONResponse({
        "ok": True,
        "stage": _lifecycle_stage_payload(stage_key, overrides, catalog),
    })


def _normalise_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _split_tags(raw: Any) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, list):
        values = raw
    else:
        values = re.split(r"[,;|]", str(raw))
    return [str(value).strip() for value in values if str(value).strip()]


def _merge_unique(*groups: list[str]) -> list[str]:
    merged: list[str] = []
    for group in groups:
        for value in group or []:
            text = str(value or "").strip()
            if text and text not in merged:
                merged.append(text)
    return merged


def _split_name_parts(full_name: str | None) -> tuple[str, str]:
    value = str(full_name or "").strip()
    if not value:
        return "", ""
    parts = value.split()
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _parse_float(value: Any) -> float:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_datetime_like(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = f"{text[:-1]}+00:00"
    if re.search(r"[+-]\d{4}$", text):
        text = f"{text[:-5]}{text[-5:-2]}:{text[-2:]}"
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed


def _infer_seeded_journey_state(last_purchase_at: Any) -> tuple[str, str]:
    last_purchase_dt = _parse_datetime_like(last_purchase_at)
    if last_purchase_dt is None:
        return "order_confirmed", "pending"

    age_days = max(
        0.0,
        (datetime.now(timezone.utc) - last_purchase_dt.astimezone(timezone.utc)).total_seconds() / 86400,
    )
    if age_days >= 10:
        return "delivered", "delivered"
    if age_days >= 3:
        return "in_transit", "in_transit"
    return "order_confirmed", "pending"


def _infer_seeded_customer_segment(total_orders: int, total_spent_paise: int, last_purchase_at: Any) -> str:
    last_purchase_dt = _parse_datetime_like(last_purchase_at)
    age_days = 0.0
    if last_purchase_dt is not None:
        age_days = max(
            0.0,
            (datetime.now(timezone.utc) - last_purchase_dt.astimezone(timezone.utc)).total_seconds() / 86400,
        )

    spent_inr = total_spent_paise / 100.0
    if spent_inr >= 5000 and total_orders >= 5 and age_days < 30:
        return "vip"
    if total_orders >= 2 and age_days < 60:
        return "repeat_buyer"
    if age_days > 120 and total_orders < 2:
        return "dormant"
    if age_days > 90:
        return "at_risk"
    return "new"


def _bootstrap_journey_customers_from_audience(connection: Any) -> int:
    scope_sql, scope_params = _journey_scope("jc")
    existing = connection.execute(
        f"SELECT COUNT(*) AS n FROM journey_customers jc WHERE {scope_sql}",
        scope_params,
    ).fetchone()["n"]
    if existing:
        return 0

    shop_domain = (getattr(settings, "default_shop_domain", "") or "").strip() or "default"
    source_rows = _safe_fetchall(
        connection,
        """
        SELECT
            id,
            email,
            name,
            first_name,
            last_name,
            phone,
            whatsapp_number,
            whatsapp_opted_in,
            total_orders,
            total_spent,
            last_order_date,
            lead_score,
            created_at,
            updated_at,
            purchase_status
        FROM customers
        WHERE deleted_at IS NULL
          AND COALESCE(whatsapp_number, phone, '') != ''
          AND (COALESCE(total_orders, 0) > 0 OR purchase_status = 'purchased')
        ORDER BY COALESCE(last_order_date, updated_at, created_at) DESC
        LIMIT 5000
        """,
    )

    inserted = 0
    now_ts = _now_iso()
    for row in source_rows:
        data = dict(row)
        phone = _normalise_phone(data.get("whatsapp_number") or data.get("phone"))
        if not phone:
            continue

        purchase_count = int(data.get("total_orders") or 0)
        total_spent_paise = int(round(float(data.get("total_spent") or 0) * 100))
        last_purchase_at = data.get("last_order_date") or data.get("updated_at") or data.get("created_at") or now_ts
        journey_stage, delivery_status = _infer_seeded_journey_state(last_purchase_at)
        customer_segment = _infer_seeded_customer_segment(purchase_count, total_spent_paise, last_purchase_at)
        full_name = data.get("name") or " ".join(
            part for part in [data.get("first_name"), data.get("last_name")] if part
        ).strip()
        order_value_paise = int(round(total_spent_paise / max(purchase_count, 1))) if purchase_count else total_spent_paise

        connection.execute(
            """
            INSERT INTO journey_customers (
                id, shop_domain, phone, name, email,
                shopify_customer_id, shopify_order_id,
                delivery_status, order_value_paise, journey_stage,
                journey_started_at, engagement_score, is_responsive,
                do_not_message, whatsapp_subscription_status,
                google_review_status, customer_segment,
                total_spent_paise, purchase_count, last_purchase_at,
                created_at, updated_at, email_address
            ) VALUES (
                ?, ?, ?, ?, ?,
                ?, ?,
                ?, ?, ?,
                ?, ?, 0,
                0, ?,
                'not_submitted', ?,
                ?, ?, ?,
                ?, ?, ?
            )
            """,
            (
                str(uuid.uuid4()),
                shop_domain,
                phone,
                full_name or phone,
                _display_email(data.get("email")),
                data.get("id"),
                None,
                delivery_status,
                order_value_paise,
                journey_stage,
                last_purchase_at,
                float(data.get("lead_score") or 0),
                "subscribed" if data.get("whatsapp_opted_in") else "unsubscribed",
                customer_segment,
                total_spent_paise,
                purchase_count,
                last_purchase_at,
                data.get("created_at") or now_ts,
                data.get("updated_at") or now_ts,
                data.get("email"),
            ),
        )
        inserted += 1

    return inserted


def _pick_latest_text(*values: Any) -> str | None:
    fallback: str | None = None
    best_value: str | None = None
    best_dt: datetime | None = None
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        fallback = fallback or text
        parsed = _parse_datetime_like(text)
        if parsed is not None and (best_dt is None or parsed > best_dt):
            best_dt = parsed
            best_value = text
    return best_value or fallback


def _pick_earliest_text(*values: Any) -> str | None:
    fallback: str | None = None
    best_value: str | None = None
    best_dt: datetime | None = None
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        fallback = fallback or text
        parsed = _parse_datetime_like(text)
        if parsed is not None and (best_dt is None or parsed < best_dt):
            best_dt = parsed
            best_value = text
    return best_value or fallback


def _audience_match_keys(email: str | None, phone: str | None) -> list[str]:
    keys: list[str] = []
    normalized_phone = _normalise_phone(phone)
    if normalized_phone:
        digits = re.sub(r"\D", "", normalized_phone)
        if digits:
            keys.append(f"phone:{digits}")

    normalized_email = (email or "").strip().lower()
    if normalized_email and not normalized_email.endswith("@whatsapp.local"):
        keys.append(f"email:{normalized_email}")
    return keys


def _is_whatsapp_lead_source(source: str | None, provider: str | None = None) -> bool:
    blocked_sources = {
        "whatsapp_csv_import",
        "whatsapp_xlsx_import",
        "whatsapp_xml_import",
        "shopify_whatsapp_import",
    }
    normalized_source = _normalise_header(source)
    normalized_provider = _normalise_header(provider)
    if normalized_source in blocked_sources or normalized_provider in blocked_sources:
        return False

    haystack = " ".join(
        part.strip().lower()
        for part in [str(source or ""), str(provider or "")]
        if part and str(part).strip()
    )
    if not haystack:
        return False
    return any(
        marker in haystack
        for marker in (
            "whatsapp",
            "wabis",
            "whatsapp_api",
            "whatsapp api",
            "wa_api",
            "wa api",
            "meta_wa",
            "meta wa",
            "cloud_api_wa",
            "cloud api wa",
        )
    )


def _looks_like_shopify_order_export(rows: list[dict[str, Any]]) -> bool:
    if not rows:
        return False
    headers = {_normalise_header(key) for key in rows[0].keys() if key is not None}
    return "name" in headers and (
        "financial_status" in headers
        or "billing_phone" in headers
        or "shipping_phone" in headers
        or "paid_at" in headers
    )


def _extract_contact_payload(row: dict[str, Any]) -> dict[str, Any]:
    normalized = {
        _normalise_header(key): ("" if value is None else str(value).strip())
        for key, value in row.items()
        if key is not None
    }

    def pick(*aliases: str) -> str:
        for alias in aliases:
            value = normalized.get(alias, "")
            if value:
                return value
        return ""

    first_name = pick("first_name", "firstname", "given_name")
    last_name = pick("last_name", "lastname", "surname", "family_name")
    name = pick("full_name", "customer_name", "billing_name", "shipping_name", "name")
    if not name:
        name = " ".join(part for part in [first_name, last_name] if part).strip()
    if name and not first_name:
        first_name, derived_last_name = _split_name_parts(name)
        last_name = last_name or derived_last_name

    source_label = pick("source", "lead_source")
    provider_label = pick("provider")
    financial_status = _normalise_header(pick("financial_status", "payment_status"))
    cancelled_at = pick("cancelled_at")
    order_key = pick("order_number", "order_id", "receipt_number")
    order_export_hints = [
        financial_status,
        pick("paid_at"),
        pick("payment_method"),
        pick("total"),
        pick("subtotal"),
        pick("billing_phone"),
        pick("shipping_phone"),
        cancelled_at,
        pick("id"),
    ]
    if not order_key and any(order_export_hints):
        order_key = pick("name", "id")

    lead_types: list[str] = []
    is_explicitly_abandoned = bool(cancelled_at) or financial_status in {"voided", "cancelled", "abandoned"}
    is_order_record = bool(order_key and any(order_export_hints))
    if is_order_record:
        lead_types.append("abandoned" if is_explicitly_abandoned else "purchased")
    if _is_whatsapp_lead_source(source_label, provider_label):
        lead_types.append("whatsapp_lead")

    tags = _split_tags(pick("tags", "tag", "labels"))
    if is_order_record:
        tags = _merge_unique(tags, ["shopify_order_export"], lead_types)

    return {
        "email": pick("email", "email_address", "email_id"),
        "phone": pick(
            "phone",
            "billing_phone",
            "shipping_phone",
            "mobile",
            "mobile_number",
            "whatsapp",
            "whatsapp_number",
            "contact_number",
        ),
        "first_name": first_name,
        "last_name": last_name,
        "name": name,
        "tags": tags,
        "lead_types": lead_types,
        "source_label": source_label or None,
        "provider_label": provider_label or None,
        "order_key": order_key or None,
        "is_order_record": is_order_record,
        "total_orders": 1 if is_order_record and "purchased" in lead_types else 0,
        "total_spent": _parse_float(pick("total", "order_total", "amount")) if is_order_record and "purchased" in lead_types else 0.0,
        "first_order_date": pick("paid_at", "created_at") or None,
        "last_order_date": pick("paid_at", "created_at") or None,
    }


def _prepare_import_payloads(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    payloads = [_extract_contact_payload(row) for row in rows]
    if not _looks_like_shopify_order_export(rows):
        return payloads

    aggregated: dict[str, dict[str, Any]] = {}
    passthrough: list[dict[str, Any]] = []
    for payload in payloads:
        identity = ((payload.get("phone") or "").strip() or (payload.get("email") or "").strip().lower())
        if not identity:
            passthrough.append(payload)
            continue

        item = aggregated.get(identity)
        if item is None:
            item = {
                "email": payload.get("email"),
                "phone": payload.get("phone"),
                "first_name": payload.get("first_name"),
                "last_name": payload.get("last_name"),
                "name": payload.get("name"),
                "tags": list(payload.get("tags") or []),
                "lead_types": list(payload.get("lead_types") or []),
                "source_label": payload.get("source_label"),
                "provider_label": payload.get("provider_label"),
                "total_orders": 0,
                "total_spent": 0.0,
                "first_order_date": None,
                "last_order_date": None,
                "_purchased_order_keys": set(),
                "_abandoned_order_keys": set(),
            }
            aggregated[identity] = item

        for field in ("email", "phone", "first_name", "last_name", "name", "source_label", "provider_label"):
            if not item.get(field) and payload.get(field):
                item[field] = payload.get(field)

        item["tags"] = _merge_unique(item.get("tags") or [], payload.get("tags") or [])
        item["lead_types"] = _merge_unique(item.get("lead_types") or [], payload.get("lead_types") or [])

        order_key = str(payload.get("order_key") or "").strip()
        if not payload.get("is_order_record") or not order_key:
            continue

        if "purchased" in (payload.get("lead_types") or []):
            if order_key not in item["_purchased_order_keys"]:
                item["_purchased_order_keys"].add(order_key)
                item["total_orders"] += 1
                item["total_spent"] += float(payload.get("total_spent") or 0.0)
            item["first_order_date"] = _pick_earliest_text(item.get("first_order_date"), payload.get("first_order_date"))
            item["last_order_date"] = _pick_latest_text(item.get("last_order_date"), payload.get("last_order_date"))
        elif "abandoned" in (payload.get("lead_types") or []):
            item["_abandoned_order_keys"].add(order_key)
            item["first_order_date"] = _pick_earliest_text(item.get("first_order_date"), payload.get("first_order_date"))
            item["last_order_date"] = _pick_latest_text(item.get("last_order_date"), payload.get("last_order_date"))

    prepared = passthrough
    for item in aggregated.values():
        item["tags"] = _merge_unique(item.get("tags") or [], item.get("lead_types") or [])
        item.pop("_purchased_order_keys", None)
        item.pop("_abandoned_order_keys", None)
        prepared.append(item)
    return prepared


def _find_existing_customer(email: str, phone: str) -> dict[str, Any] | None:
    with get_db_connection() as conn:
        row = None
        if email:
            row = conn.execute(
                "SELECT * FROM customers WHERE LOWER(email) = ? AND deleted_at IS NULL LIMIT 1",
                (email.lower(),),
            ).fetchone()
        if row is None and phone:
            row = conn.execute(
                """
                SELECT *
                FROM customers
                WHERE deleted_at IS NULL AND (phone = ? OR whatsapp_number = ?)
                LIMIT 1
                """,
                (phone, phone),
            ).fetchone()
        return dict(row) if row else None


def _ensure_whatsapp_contact(payload: dict[str, Any], source: str, default_tags: list[str] | None = None) -> tuple[str, dict[str, Any] | None]:
    email = (payload.get("email") or "").strip().lower()
    phone = _normalise_phone(payload.get("phone"))
    existing = _find_existing_customer(email, phone)

    resolved_phone = phone or (existing.get("whatsapp_number") if existing else "") or (existing.get("phone") if existing else "")
    if not resolved_phone:
        return "skipped_no_phone", None

    target_email = (existing.get("email") if existing else "") or email
    if not target_email:
        phone_digits = re.sub(r"\D", "", resolved_phone) or uuid.uuid4().hex[:12]
        target_email = f"wa_{phone_digits}@whatsapp.local"

    existed_before = _find_existing_customer(target_email, resolved_phone)
    merged_tags = _merge_unique(default_tags or [], payload.get("tags") or [], payload.get("lead_types") or [])

    customer = upsert_customer(
        email=target_email,
        name=payload.get("name") or None,
        first_name=payload.get("first_name") or None,
        last_name=payload.get("last_name") or None,
        phone=resolved_phone,
        tags=merged_tags,
        source=source,
        whatsapp_opted_in=True,
        preferred_channel="whatsapp",
    )

    payload_total_orders = int(payload.get("total_orders") or 0)
    if customer and payload_total_orders > 0:
        sync_order_stats(
            customer["email"],
            total_orders=max(int(customer.get("total_orders") or 0), payload_total_orders),
            total_spent=max(float(customer.get("total_spent") or 0), float(payload.get("total_spent") or 0.0)),
            last_order_date=_pick_latest_text(customer.get("last_order_date"), payload.get("last_order_date")),
            first_order_date=_pick_earliest_text(customer.get("first_order_date"), payload.get("first_order_date")),
        )

    return ("updated" if existed_before else "created"), customer


def _import_contacts(rows: list[dict[str, Any]], source: str, default_tags: list[str] | None = None) -> dict[str, Any]:
    prepared_rows = _prepare_import_payloads(rows)
    created = 0
    updated = 0
    skipped = 0
    skipped_no_phone = 0
    errors: list[dict[str, Any]] = []

    for index, payload in enumerate(prepared_rows, start=1):
        try:
            outcome, _customer = _ensure_whatsapp_contact(payload, source=source, default_tags=default_tags)
            if outcome == "created":
                created += 1
            elif outcome == "updated":
                updated += 1
            else:
                skipped += 1
                if outcome == "skipped_no_phone":
                    skipped_no_phone += 1
        except Exception as exc:
            errors.append({
                "row": index,
                "email": payload.get("email") or None,
                "phone": payload.get("phone") or None,
                "error": str(exc),
            })

    return {
        "ok": True,
        "total_rows": len(rows),
        "normalized_contacts": len(prepared_rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "skipped_no_phone": skipped_no_phone,
        "errors": errors[:20],
    }


def _merge_shopify_contacts(customer_rows: list[dict[str, Any]], order_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}

    def merge_row(row: dict[str, Any]) -> None:
        payload = _extract_contact_payload(row)
        phone = _normalise_phone(payload.get("phone"))
        key = (payload.get("email") or phone or f"shopify:{row.get('id')}").strip().lower()
        if not key:
            return
        existing = merged.get(key)
        if existing is None:
            merged[key] = dict(row)
            return

        for field in ("email", "phone", "first_name", "last_name", "name"):
            if not (existing.get(field) or "") and (row.get(field) or ""):
                existing[field] = row.get(field)

        existing_tags = _split_tags(existing.get("tags"))
        next_tags = _split_tags(row.get("tags"))
        if existing_tags or next_tags:
            existing["tags"] = list(dict.fromkeys([*existing_tags, *next_tags]))

    for row in customer_rows:
        merge_row(row)
    for row in order_rows:
        merge_row(row)
    return list(merged.values())


def _read_csv_rows(raw: bytes) -> list[dict[str, Any]]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader if any((value or "").strip() for value in row.values())]


def _read_xml_rows(raw: bytes) -> list[dict[str, Any]]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")

    root = ET.fromstring(text)
    children = list(root)
    if not children:
        return []
    if all(not list(child) for child in children):
        return [{child.tag: child.text or "" for child in children}]

    records: list[dict[str, Any]] = []
    for child in children:
        grand_children = list(child)
        if not grand_children:
            continue
        records.append({grand.tag: grand.text or "" for grand in grand_children})
    return records


def _read_xlsx_rows(raw: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError as exc:
        raise RuntimeError("XLSX import requires openpyxl on the server") from exc

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        return []

    normalized_headers = [str(header).strip() if header is not None else "" for header in headers]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        if not values or not any(value not in (None, "") for value in values):
            continue
        row: dict[str, Any] = {}
        for idx, header in enumerate(normalized_headers):
            if not header:
                continue
            row[header] = values[idx] if idx < len(values) else ""
        rows.append(row)
    return rows


def _primary_audience_lead_type(lead_types: list[str]) -> str:
    for bucket in _AUDIENCE_BUCKET_PRIORITY:
        if bucket in lead_types:
            return bucket
    return "general"


def _derive_customer_lead_types(row: dict[str, Any]) -> list[str]:
    tags = {
        _normalise_header(tag)
        for tag in _json_loads(row.get("tags"), [])
        if str(tag or "").strip()
    }
    status_tokens = {
        _normalise_header(row.get("customer_status") or ""),
        _normalise_header(row.get("purchase_status") or ""),
        _normalise_header(row.get("engagement_label") or ""),
    }
    lead_types: list[str] = []
    if int(row.get("total_orders") or 0) > 0 or str(row.get("purchase_status") or "").strip().lower() == "purchased" or "purchased" in tags:
        lead_types.append("purchased")
    if "interested" in tags or "interested" in status_tokens or "purchase_intent" in status_tokens or "intent" in tags:
        lead_types.append("interested")
    if "abandoned" in tags or "abandoned_order" in tags:
        lead_types.append("abandoned")
    if "whatsapp_lead" in tags or "whatsapplead" in tags or _is_whatsapp_lead_source(row.get("source"), None):
        lead_types.append("whatsapp_lead")
    if "promotional_lead" in tags or "promotional_customer" in tags or "promo_lead" in tags:
        lead_types.append("promotional_lead")
    return lead_types


def _format_whatsapp_customer(row: dict[str, Any]) -> dict[str, Any]:
    lead_types = _derive_customer_lead_types(row)
    return {
        "id": row.get("id"),
        "customer_uid": row.get("customer_uid"),
        "email": _display_email(row.get("email")),
        "phone": row.get("whatsapp_number") or row.get("phone"),
        "full_name": row.get("name") or " ".join(
            part for part in [row.get("first_name"), row.get("last_name")] if part
        ).strip() or None,
        "first_name": row.get("first_name"),
        "last_name": row.get("last_name"),
        "tags": _json_loads(row.get("tags"), []),
        "source": row.get("source"),
        "customer_status": row.get("customer_status"),
        "purchase_status": row.get("purchase_status"),
        "whatsapp_opted_in": bool(row.get("whatsapp_opted_in")),
        "total_orders": row.get("total_orders") or 0,
        "total_spent": round(float(row.get("total_spent") or 0), 2),
        "last_order_date": row.get("last_order_date"),
        "engagement_label": row.get("engagement_label") or "cold",
        "preferred_channel": row.get("preferred_channel") or "auto",
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "lead_score": int(row.get("lead_score") or 0),
        "lead_type": _primary_audience_lead_type(lead_types),
        "lead_types": lead_types,
        "source_types": [row.get("source")] if row.get("source") else [],
    }


def _build_external_audience_record(
    *,
    record_id: str,
    email: str | None,
    phone: str | None,
    full_name: str | None,
    first_name: str | None = None,
    last_name: str | None = None,
    tags: list[str] | None = None,
    source: str,
    updated_at: str | None,
    lead_types: list[str],
    whatsapp_opted_in: bool = False,
    customer_status: str = "lead",
    engagement_label: str = "cold",
    lead_score: int = 0,
) -> dict[str, Any]:
    if full_name and (not first_name and not last_name):
        first_name, last_name = _split_name_parts(full_name)
    return {
        "id": record_id,
        "customer_uid": None,
        "email": _display_email(email),
        "phone": _normalise_phone(phone),
        "full_name": full_name,
        "first_name": first_name,
        "last_name": last_name,
        "tags": _merge_unique(tags or [], lead_types),
        "source": source,
        "source_types": [source],
        "customer_status": customer_status,
        "purchase_status": "purchased" if "purchased" in lead_types else "lead",
        "whatsapp_opted_in": bool(whatsapp_opted_in),
        "total_orders": 0,
        "total_spent": 0.0,
        "last_order_date": None,
        "engagement_label": engagement_label,
        "preferred_channel": "whatsapp",
        "created_at": updated_at,
        "updated_at": updated_at,
        "lead_score": int(lead_score or 0),
        "lead_type": _primary_audience_lead_type(lead_types),
        "lead_types": _merge_unique(lead_types),
    }


def _merge_audience_record(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    for field in (
        "customer_uid",
        "email",
        "phone",
        "full_name",
        "first_name",
        "last_name",
        "customer_status",
        "purchase_status",
        "engagement_label",
        "preferred_channel",
        "created_at",
    ):
        if not existing.get(field) and incoming.get(field):
            existing[field] = incoming[field]

    if incoming.get("customer_status") and existing.get("customer_status") in {None, "", "new"}:
        existing["customer_status"] = incoming["customer_status"]
    if existing.get("source") in {None, "", "manual"} and incoming.get("source"):
        existing["source"] = incoming["source"]

    existing["tags"] = _merge_unique(existing.get("tags") or [], incoming.get("tags") or [])
    existing["lead_types"] = _merge_unique(existing.get("lead_types") or [], incoming.get("lead_types") or [])
    existing["source_types"] = _merge_unique(existing.get("source_types") or [], incoming.get("source_types") or [])
    existing["whatsapp_opted_in"] = bool(existing.get("whatsapp_opted_in") or incoming.get("whatsapp_opted_in"))
    existing["total_orders"] = max(int(existing.get("total_orders") or 0), int(incoming.get("total_orders") or 0))
    existing["total_spent"] = max(float(existing.get("total_spent") or 0), float(incoming.get("total_spent") or 0))
    existing["lead_score"] = max(int(existing.get("lead_score") or 0), int(incoming.get("lead_score") or 0))
    existing["last_order_date"] = _pick_latest_text(existing.get("last_order_date"), incoming.get("last_order_date"))
    existing["updated_at"] = _pick_latest_text(existing.get("updated_at"), incoming.get("updated_at"))
    existing["lead_type"] = _primary_audience_lead_type(existing.get("lead_types") or [])
    return existing


def _add_audience_record(
    records: list[dict[str, Any]],
    index: dict[str, dict[str, Any]],
    incoming: dict[str, Any],
) -> None:
    keys = _audience_match_keys(incoming.get("email"), incoming.get("phone"))
    if not keys:
        return

    target = next((index[key] for key in keys if key in index), None)
    if target is None:
        target = dict(incoming)
        records.append(target)
    else:
        _merge_audience_record(target, incoming)

    for key in keys:
        index[key] = target


def _safe_fetchall(connection: Any, sql: str, params: list[Any] | tuple[Any, ...] | None = None) -> list[Any]:
    try:
        return connection.execute(sql, params or []).fetchall()
    except Exception as exc:
        logger.debug("Skipping WhatsApp audience source query: %s", exc)
        return []


def _load_whatsapp_audience(connection: Any) -> tuple[list[dict[str, Any]], dict[str, int]]:
    records: list[dict[str, Any]] = []
    index: dict[str, dict[str, Any]] = {}

    customer_rows = _safe_fetchall(
        connection,
        """
        SELECT *
        FROM customers
        WHERE deleted_at IS NULL AND COALESCE(whatsapp_number, phone, '') != ''
        """,
    )
    for row in customer_rows:
        _add_audience_record(records, index, _format_whatsapp_customer(dict(row)))

    promo_rows = _safe_fetchall(
        connection,
        """
        SELECT id, email, first_name, last_name, phone, tags, status, subscribed_to_promo, created_at, updated_at
        FROM promotional_customers
        WHERE COALESCE(phone, '') != ''
        """,
    )
    for row in promo_rows:
        data = dict(row)
        full_name = " ".join(part for part in [data.get("first_name"), data.get("last_name")] if part).strip() or None
        _add_audience_record(
            records,
            index,
            _build_external_audience_record(
                record_id=f"promo:{data.get('id')}",
                email=data.get("email"),
                phone=data.get("phone"),
                full_name=full_name,
                first_name=data.get("first_name"),
                last_name=data.get("last_name"),
                tags=_split_tags(data.get("tags")),
                source="promotional_customers",
                updated_at=data.get("updated_at") or data.get("created_at"),
                lead_types=["promotional_lead"],
                whatsapp_opted_in=bool(data.get("subscribed_to_promo")),
                customer_status=data.get("status") or "lead",
                engagement_label="warm" if data.get("subscribed_to_promo") else "cold",
            ),
        )

    abandoned_rows = _safe_fetchall(
        connection,
        """
        SELECT id, name, email, phone, interest_product, interest_category, engagement_score, created_at, updated_at
        FROM abandoned_leads
        WHERE COALESCE(phone, '') != ''
        """,
    )
    for row in abandoned_rows:
        data = dict(row)
        abandoned_tags = ["abandoned"]
        if data.get("interest_category"):
            abandoned_tags.append(_normalise_header(data["interest_category"]))
        if data.get("interest_product"):
            abandoned_tags.append(_normalise_header(data["interest_product"]))
        _add_audience_record(
            records,
            index,
            _build_external_audience_record(
                record_id=f"abandoned:{data.get('id')}",
                email=data.get("email"),
                phone=data.get("phone"),
                full_name=data.get("name"),
                tags=abandoned_tags,
                source="abandoned_leads",
                updated_at=data.get("updated_at") or data.get("created_at"),
                lead_types=["abandoned"],
                customer_status="abandoned",
                engagement_label="warm" if float(data.get("engagement_score") or 0) > 0 else "cold",
            ),
        )

    lead_rows = _safe_fetchall(
        connection,
        """
        SELECT id, name, email, phone, source, provider, consent, created_at, captured_at
        FROM leads
        WHERE COALESCE(phone, '') != ''
        """,
    )
    for row in lead_rows:
        data = dict(row)
        if not _is_whatsapp_lead_source(data.get("source"), data.get("provider")):
            continue
        lead_tags = _merge_unique(
            ["whatsapp_lead"],
            [
                _normalise_header(data.get("provider")) if data.get("provider") else "",
                _normalise_header(data.get("source")) if data.get("source") else "",
            ],
        )
        _add_audience_record(
            records,
            index,
            _build_external_audience_record(
                record_id=f"lead:{data.get('id')}",
                email=data.get("email"),
                phone=data.get("phone"),
                full_name=data.get("name"),
                tags=lead_tags,
                source=data.get("provider") or data.get("source") or "whatsapp_lead",
                updated_at=data.get("captured_at") or data.get("created_at"),
                lead_types=["whatsapp_lead"],
                whatsapp_opted_in=bool(data.get("consent")),
                customer_status="lead",
                engagement_label="warm" if data.get("consent") else "cold",
            ),
        )

    for record in records:
        record["tags"] = _merge_unique(record.get("tags") or [])
        record["lead_types"] = [bucket for bucket in _AUDIENCE_BUCKET_PRIORITY if bucket in (record.get("lead_types") or [])]
        record["lead_type"] = _primary_audience_lead_type(record.get("lead_types") or [])
        record["source_types"] = _merge_unique(record.get("source_types") or [])
        record["full_name"] = record.get("full_name") or " ".join(
            part for part in [record.get("first_name"), record.get("last_name")] if part
        ).strip() or None
        record["phone"] = _normalise_phone(record.get("phone"))
        record["total_orders"] = int(record.get("total_orders") or 0)
        record["total_spent"] = round(float(record.get("total_spent") or 0), 2)

    records.sort(
        key=lambda item: (
            (_parse_datetime_like(item.get("updated_at")) or datetime.fromtimestamp(0, tz=timezone.utc)).timestamp(),
            str(item.get("id") or ""),
        ),
        reverse=True,
    )

    summary = {
        "total": len(records),
        "opted_in": sum(1 for item in records if item.get("whatsapp_opted_in")),
        "purchased": 0,
        "interested": 0,
        "abandoned": 0,
        "whatsapp_leads": 0,
        "promotional_leads": 0,
    }
    for record in records:
        lead_types = set(record.get("lead_types") or [])
        for bucket, summary_key in _AUDIENCE_SUMMARY_KEY_MAP.items():
            if bucket in lead_types:
                summary[summary_key] += 1

    return records, summary


def _ensure_sales_os_tables(conn: Any) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS retargeting_sync_runs (
            id              TEXT PRIMARY KEY,
            audience_key    TEXT NOT NULL,
            audience_name   TEXT NOT NULL,
            platform        TEXT NOT NULL,
            filter_json     TEXT NOT NULL DEFAULT '{}',
            matched_count   INTEGER NOT NULL DEFAULT 0,
            status          TEXT NOT NULL DEFAULT 'pending',
            result_json     TEXT NOT NULL DEFAULT '{}',
            created_at      TEXT NOT NULL,
            updated_at      TEXT NOT NULL
        )
        """
    )
    conn.execute('CREATE INDEX IF NOT EXISTS idx_retargeting_sync_runs_platform ON retargeting_sync_runs(platform, created_at DESC)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_retargeting_sync_runs_audience ON retargeting_sync_runs(audience_key, platform, created_at DESC)')
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS order_lifecycle_audit (
            id                  TEXT PRIMARY KEY,
            order_id            TEXT,
            customer_id         TEXT,
            phone               TEXT,
            event_type          TEXT NOT NULL,
            lifecycle_status    TEXT NOT NULL,
            template_name       TEXT,
            message_id          TEXT,
            error_detail        TEXT,
            payload_json        TEXT NOT NULL DEFAULT '{}',
            created_at          TEXT NOT NULL
        )
        """
    )
    conn.execute('CREATE INDEX IF NOT EXISTS idx_order_lifecycle_audit_order ON order_lifecycle_audit(order_id, created_at DESC)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_order_lifecycle_audit_status ON order_lifecycle_audit(lifecycle_status, created_at DESC)')


def _audience_label_for_key(audience_key: str) -> str:
    labels = {
        "all": "All WhatsApp Customers",
        "purchased": "Purchased Customers",
        "interested": "Interested Non-purchased Customers",
        "non_purchased": "Non-purchased WhatsApp Leads",
        "abandoned": "Abandoned / Open Order Prospects",
        "whatsapp_leads": "WhatsApp Leads",
        "hot_leads": "Hot Leads Score 80+",
        "review_pending": "Delivered Customers Without Review",
    }
    return labels.get(audience_key, audience_key.replace("_", " ").title())


def _row_matches_sales_audience(row: dict[str, Any], audience_key: str) -> bool:
    lead_types = {str(item).lower() for item in (row.get("lead_types") or [])}
    score = int(row.get("lead_score") or 0)
    total_orders = int(row.get("total_orders") or 0)
    if audience_key == "purchased":
        return "purchased" in lead_types or total_orders > 0
    if audience_key == "interested":
        labels = {str(tag).lower() for tag in (row.get("tags") or [])}
        statuses = {
            str(row.get("customer_status") or "").lower(),
            str(row.get("purchase_status") or "").lower(),
            str(row.get("engagement_label") or "").lower(),
        }
        return total_orders == 0 and ("interested" in lead_types or "interested" in labels or "interested" in statuses or score >= 70)
    if audience_key == "non_purchased":
        return "purchased" not in lead_types and total_orders == 0
    if audience_key == "abandoned":
        return "abandoned" in lead_types
    if audience_key == "whatsapp_leads":
        return "whatsapp_lead" in lead_types
    if audience_key == "hot_leads":
        return score >= 80
    if audience_key in {"all", ""}:
        return True
    return audience_key in lead_types


def _sales_audience_counts(records: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "all": len(records),
        "purchased": sum(1 for row in records if _row_matches_sales_audience(row, "purchased")),
        "interested": sum(1 for row in records if _row_matches_sales_audience(row, "interested")),
        "non_purchased": sum(1 for row in records if _row_matches_sales_audience(row, "non_purchased")),
        "abandoned": sum(1 for row in records if _row_matches_sales_audience(row, "abandoned")),
        "whatsapp_leads": sum(1 for row in records if _row_matches_sales_audience(row, "whatsapp_leads")),
        "hot_leads": sum(1 for row in records if _row_matches_sales_audience(row, "hot_leads")),
    }


def _record_lifecycle_audit(
    conn: Any,
    *,
    event_type: str,
    lifecycle_status: str,
    order_id: str | None = None,
    customer_id: str | None = None,
    phone: str | None = None,
    template_name: str | None = None,
    message_id: str | None = None,
    error_detail: str | None = None,
    payload: dict[str, Any] | None = None,
) -> None:
    _ensure_sales_os_tables(conn)
    conn.execute(
        """
        INSERT INTO order_lifecycle_audit
            (id, order_id, customer_id, phone, event_type, lifecycle_status,
             template_name, message_id, error_detail, payload_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            str(uuid.uuid4()),
            order_id,
            customer_id,
            phone,
            event_type,
            lifecycle_status,
            template_name,
            message_id,
            error_detail,
            json.dumps(payload or {}, ensure_ascii=True),
            _now_iso(),
        ),
    )


def _matches_audience_search(record: dict[str, Any], search: str) -> bool:
    needle = search.strip().lower()
    if not needle:
        return True
    phone_needle = re.sub(r"\D", "", search)
    if phone_needle:
        phone_digits = re.sub(r"\D", "", record.get("phone") or "")
        if phone_needle in phone_digits:
            return True

    haystacks = [
        record.get("full_name"),
        record.get("email"),
        record.get("source"),
        " ".join(record.get("tags") or []),
        " ".join(record.get("lead_types") or []),
        " ".join(record.get("source_types") or []),
    ]
    return any(needle in str(value or "").lower() for value in haystacks)


def _record_status_event(
    *,
    message_id: str,
    recipient_phone: str,
    status: str,
    template_name: str,
    raw_payload: dict[str, Any],
    error_code: str | None = None,
    error_title: str | None = None,
    error_detail: str | None = None,
) -> None:
    with get_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO whatsapp_message_status_events
                (id, message_id, recipient_phone, status, template_name,
                 error_code, error_title, error_detail, raw_payload_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(uuid.uuid4()),
                message_id,
                recipient_phone,
                status,
                template_name,
                error_code,
                error_title,
                error_detail,
                json.dumps(raw_payload, ensure_ascii=True),
                _now_iso(),
            ),
        )


def _meta_template_components(template_name: str, language_code: str) -> dict[str, Any] | None:
    templates = _get_cached_templates("meta")
    normalized_language = (language_code or "").lower()

    for template in templates:
        if template.get("name") != template_name:
            continue
        template_language = str(template.get("locale") or "").lower()
        if normalized_language and template_language and template_language != normalized_language:
            continue
        return template

    for template in templates:
        if template.get("name") == template_name:
            return template
    return None


def _build_meta_send_components(template: dict[str, Any] | None, params: list[str]) -> list[dict[str, Any]]:
    if not template or not params:
        return []

    values = [str(value) for value in params]
    cursor = 0
    components: list[dict[str, Any]] = []

    def take_values(count: int) -> list[str]:
        nonlocal cursor
        if count <= 0:
            return []
        next_values = values[cursor:cursor + count]
        cursor += count
        return next_values

    header = template.get("header") or {}
    header_values = take_values(int(header.get("vars") or 0))
    if header_values:
        components.append({
            "type": "header",
            "parameters": [{"type": "text", "text": value} for value in header_values],
        })

    body = template.get("body") or {}
    body_values = take_values(int(body.get("vars") or 0))
    if body_values:
        components.append({
            "type": "body",
            "parameters": [{"type": "text", "text": value} for value in body_values],
        })

    for button_index, button in enumerate(template.get("buttons") or []):
        button_values = take_values(int(button.get("vars") or 0))
        if not button_values:
            continue
        components.append({
            "type": "button",
            "sub_type": str(button.get("sub_type") or button.get("type") or "URL").lower(),
            "index": str(button_index),
            "parameters": [{"type": "text", "text": value} for value in button_values],
        })

    return components


# ─── Endpoints ───────────────────────────────────────────────────────────────

@router.get("/crm/whatsapp/templates")
def get_whatsapp_templates(force: bool = False, refresh: bool = False) -> JSONResponse:
    """
    Return normalised templates from Wabis and Meta.
    Cached for 30 min; pass ?force=true to bust cache.
    """
    now = time.time()

    cache_bust = force or refresh

    if not cache_bust:
        wabis_cached = _cache.get("wabis")
        meta_cached = _cache.get("meta")
        if wabis_cached and (now - wabis_cached[1]) < _CACHE_TTL:
            wabis = wabis_cached[0]
        else:
            wabis = _fetch_wabis_templates()
            _cache["wabis"] = (wabis, now)

        if meta_cached and (now - meta_cached[1]) < _CACHE_TTL:
            meta_list = meta_cached[0]
        else:
            meta_list = _fetch_meta_templates()
            _cache["meta"] = (meta_list, now)
    else:
        wabis = _fetch_wabis_templates()
        _cache["wabis"] = (wabis, now)
        meta_list = _fetch_meta_templates()
        _cache["meta"] = (meta_list, now)

    meta_configured = bool(
        settings.meta_access_token
        and settings.meta_waba_id
    )

    return JSONResponse({
        "wabis": wabis,
        "meta": meta_list,
        "meta_configured": meta_configured,
        "wabis_phone_number_id": settings.wabis_phone_number_id or None,
        "meta_waba_id": settings.meta_waba_id or None,
        "meta_phone_number_id": settings.meta_phone_number_id or None,
        "cached_at": int(now),
    })


@router.get("/crm/whatsapp/dashboard")
def whatsapp_dashboard() -> JSONResponse:
    scope_sql, scope_params = _journey_scope("jc")
    cutoff_30d = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    wabis_templates = _get_cached_templates("wabis")
    meta_templates = _get_cached_templates("meta")

    with get_db_connection() as conn:
        _bootstrap_journey_customers_from_audience(conn)
        _audience_rows, audience_summary = _load_whatsapp_audience(conn)

        journey_total = conn.execute(
            f"SELECT COUNT(*) AS n FROM journey_customers jc WHERE {scope_sql}",
            scope_params,
        ).fetchone()["n"]
        active_sessions = conn.execute(
            "SELECT COUNT(*) AS n FROM conversation_sessions WHERE is_active = 1"
            + (" AND shop_domain = ?" if scope_params else ""),
            scope_params,
        ).fetchone()["n"]

        message_summary = conn.execute(
            """
            WITH latest AS (
                SELECT
                    status,
                    message_id,
                    created_at,
                    ROW_NUMBER() OVER (PARTITION BY message_id ORDER BY created_at DESC) AS rn
                FROM whatsapp_message_status_events
                WHERE created_at >= ?
            )
            SELECT
                COUNT(*) AS total_messages,
                SUM(CASE WHEN status IN ('sent', 'accepted') THEN 1 ELSE 0 END) AS sent_count,
                SUM(CASE WHEN status = 'delivered' THEN 1 ELSE 0 END) AS delivered_count,
                SUM(CASE WHEN status = 'read' THEN 1 ELSE 0 END) AS read_count,
                SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) AS failed_count
            FROM latest
            WHERE rn = 1
            """,
            (cutoff_30d,),
        ).fetchone()

        stage_rows = conn.execute(
            f"""
            SELECT journey_stage, COUNT(*) AS cnt
            FROM journey_customers jc
            WHERE {scope_sql}
            GROUP BY journey_stage
            ORDER BY cnt DESC, journey_stage ASC
            """,
            scope_params,
        ).fetchall()
        segment_rows = conn.execute(
            f"""
            SELECT customer_segment, COUNT(*) AS cnt
            FROM journey_customers jc
            WHERE {scope_sql}
            GROUP BY customer_segment
            ORDER BY cnt DESC, customer_segment ASC
            """,
            scope_params,
        ).fetchall()
        engagement_rows = conn.execute(
            f"""
            SELECT je.event_type, COUNT(*) AS cnt
            FROM journey_engagement_events je
            JOIN journey_customers jc ON jc.id = je.customer_id
            WHERE {scope_sql} AND je.created_at >= ?
            GROUP BY je.event_type
            ORDER BY cnt DESC, je.event_type ASC
            LIMIT 8
            """,
            [*scope_params, cutoff_30d],
        ).fetchall()
        top_template_rows = conn.execute(
            """
            SELECT template_name, COUNT(*) AS cnt, MAX(created_at) AS last_seen_at
            FROM whatsapp_message_status_events
            WHERE created_at >= ? AND COALESCE(template_name, '') != ''
            GROUP BY template_name
            ORDER BY cnt DESC, last_seen_at DESC
            LIMIT 8
            """,
            (cutoff_30d,),
        ).fetchall()

        score_tier_rows = conn.execute(
            """
            SELECT
                CASE
                    WHEN COALESCE(lead_score, 0) >= 90 THEN 'buyers'
                    WHEN COALESCE(lead_score, 0) >= 80 THEN 'purchase_intent'
                    WHEN COALESCE(lead_score, 0) >= 50 THEN 'warm'
                    WHEN COALESCE(lead_score, 0) >= 10 THEN 'behavioral'
                    ELSE 'cold'
                END AS tier,
                COUNT(*) AS count
            FROM customers
            WHERE deleted_at IS NULL
            GROUP BY 1
            """,
        ).fetchall()

    return JSONResponse({
        "ok": True,
        "totals": {
            "audience_customers": audience_summary["total"],
            "opted_in_customers": audience_summary["opted_in"],
            "journey_customers": journey_total,
            "active_sessions": active_sessions,
            "messages_30d": message_summary["total_messages"] or 0,
            "sent_30d": message_summary["sent_count"] or 0,
            "delivered_30d": message_summary["delivered_count"] or 0,
            "read_30d": message_summary["read_count"] or 0,
            "failed_30d": message_summary["failed_count"] or 0,
        },
        "audience_breakdown": {
            "purchased": audience_summary["purchased"],
            "interested": audience_summary["interested"],
            "abandoned": audience_summary["abandoned"],
            "whatsapp_leads": audience_summary["whatsapp_leads"],
            "promotional_leads": audience_summary["promotional_leads"],
        },
        "template_sources": {
            "wabis": len(wabis_templates),
            "meta": len(meta_templates),
            "meta_configured": bool(settings.meta_access_token and settings.meta_waba_id),
        },
        "stages": [{"label": row["journey_stage"] or "unknown", "count": row["cnt"]} for row in stage_rows],
        "segments": [{"label": row["customer_segment"] or "unknown", "count": row["cnt"]} for row in segment_rows],
        "engagement": [{"label": row["event_type"], "count": row["cnt"]} for row in engagement_rows],
        "top_templates": [dict(row) for row in top_template_rows],
        "score_breakdown": {row["tier"]: row["count"] for row in score_tier_rows},
        "sync_interval_minutes": int(_CACHE_TTL / 60),
    })


@router.get("/crm/whatsapp/customers")
def whatsapp_customers(
    search: str = Query(default=""),
    lead_type: str = Query(default="all"),
    labels: str = Query(default=""),
    status_filter: str = Query(default="all", alias="status"),
    engagement_label: str = Query(default="all"),
    min_score: int = Query(default=0, ge=0, le=100),
    max_score: int = Query(default=100, ge=0, le=100),
    language: str = Query(default="all"),
    date_field: str = Query(default="updated_at"),
    date_from: str = Query(default=""),
    date_to: str = Query(default=""),
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> JSONResponse:
    with get_db_connection() as conn:
        audience_rows, audience_summary = _load_whatsapp_audience(conn)

    filtered_rows = [row for row in audience_rows if _matches_audience_search(row, search)]
    if lead_type != "all":
        filtered_rows = [row for row in filtered_rows if lead_type in (row.get("lead_types") or [])]
    selected_labels = [item.lower() for item in _normalize_multi_value(labels)]
    if selected_labels:
        filtered_rows = [
            row for row in filtered_rows
            if {str(tag).lower() for tag in (row.get("tags") or [])}.intersection(selected_labels)
        ]
    if status_filter != "all":
        status_value = status_filter.strip().lower()
        filtered_rows = [
            row for row in filtered_rows
            if status_value in {
                str(row.get("customer_status") or "").lower(),
                str(row.get("purchase_status") or "").lower(),
                str(row.get("engagement_label") or "").lower(),
                *[str(item).lower() for item in (row.get("lead_types") or [])],
            }
        ]
    if engagement_label != "all":
        filtered_rows = [row for row in filtered_rows if str(row.get("engagement_label") or "").lower() == engagement_label.strip().lower()]
    if min_score > 0 or max_score < 100:
        filtered_rows = [
            row for row in filtered_rows
            if min_score <= int(row.get("lead_score") or 0) <= max_score
        ]
    date_filters = _campaign_filters({"date_field": date_field, "date_from": date_from, "date_to": date_to})
    if date_filters.get("date_from") or date_filters.get("date_to"):
        filtered_rows = [row for row in filtered_rows if _campaign_row_matches(row, {**date_filters, "score_min": 0, "score_max": 100})]
    if language != "all":
        lang_tag = "MAL_Customer" if language == "mal" else "Eng_customer"
        filtered_rows = [
            row for row in filtered_rows
            if lang_tag.lower() in [str(t).lower() for t in (row.get("tags") or [])]
        ]
    total = len(filtered_rows)
    rows = filtered_rows[offset:offset + limit]

    return JSONResponse({
        "ok": True,
        "customers": rows,
        "total": total,
        "count": len(rows),
        "summary": {
            "total": audience_summary["total"] or 0,
            "opted_in": audience_summary["opted_in"] or 0,
            "purchased": audience_summary["purchased"] or 0,
            "interested": audience_summary["interested"] or 0,
            "abandoned": audience_summary["abandoned"] or 0,
            "whatsapp_leads": audience_summary["whatsapp_leads"] or 0,
            "promotional_leads": audience_summary["promotional_leads"] or 0,
        },
    })


@router.get("/crm/whatsapp/sales-os")
def whatsapp_sales_os_command_center() -> JSONResponse:
    """Return the no-new-tab Sales OS operating snapshot for dashboard, customers, and journeys."""
    now = datetime.now(timezone.utc)
    cutoff_7d = (now - timedelta(days=7)).isoformat()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    automation_start = _automation_start_iso()

    with get_db_connection() as conn:
        _ensure_sales_os_tables(conn)
        _bootstrap_journey_customers_from_audience(conn)
        audience_rows, _summary = _load_whatsapp_audience(conn)

        lifecycle = conn.execute(
            """
            SELECT
                COUNT(*) AS total_journey_orders,
                SUM(CASE WHEN COALESCE(journey_started_at, created_at) >= ? THEN 1 ELSE 0 END) AS automation_orders,
                SUM(CASE WHEN COALESCE(day1_sent, 0) = 0 AND do_not_message = 0 AND whatsapp_subscription_status = 'subscribed' AND COALESCE(journey_started_at, created_at) >= ? THEN 1 ELSE 0 END) AS order_confirmations_pending,
                SUM(CASE WHEN COALESCE(waybill_id, '') != '' AND COALESCE(day2_sent, 0) = 0 AND do_not_message = 0 AND whatsapp_subscription_status = 'subscribed' AND COALESCE(updated_at, created_at) >= ? THEN 1 ELSE 0 END) AS tracking_messages_pending,
                SUM(CASE WHEN delivery_status = 'delivered' AND COALESCE(day5_sent, 0) = 0 AND do_not_message = 0 AND whatsapp_subscription_status = 'subscribed' AND COALESCE(delivered_at, updated_at, created_at) >= ? THEN 1 ELSE 0 END) AS delivery_messages_pending,
                SUM(CASE WHEN COALESCE(waybill_id, '') = '' AND delivery_status IN ('pending', 'order_confirmed') AND COALESCE(journey_started_at, created_at) >= ? THEN 1 ELSE 0 END) AS orders_missing_tracking,
                SUM(CASE WHEN delivery_status = 'in_transit' AND COALESCE(updated_at, created_at) >= ? THEN 1 ELSE 0 END) AS in_transit_orders,
                SUM(CASE WHEN delivery_status = 'delivered' AND COALESCE(delivered_at, updated_at, created_at) >= ? THEN 1 ELSE 0 END) AS delivered_orders
            FROM journey_customers
            """,
            (automation_start, automation_start, automation_start, automation_start, automation_start, automation_start, automation_start),
        ).fetchone()

        shopify_order_counts = conn.execute(
            """
            SELECT
                COUNT(*) AS total_orders,
                SUM(CASE WHEN created_at_shopify >= ? THEN 1 ELSE 0 END) AS orders_today,
                SUM(CASE WHEN created_at_shopify >= ? THEN 1 ELSE 0 END) AS orders_7d,
                SUM(CASE WHEN fulfillment_status IN ('fulfilled', 'partial') THEN 1 ELSE 0 END) AS fulfilled_orders,
                SUM(CASE WHEN COALESCE(fulfillment_status, '') NOT IN ('fulfilled', 'partial') THEN 1 ELSE 0 END) AS unfulfilled_orders
            FROM shopify_orders
            """,
            (today_start, cutoff_7d),
        ).fetchone()

        latest_failed = conn.execute(
            """
            WITH latest AS (
                SELECT *, ROW_NUMBER() OVER (PARTITION BY message_id ORDER BY created_at DESC) AS rn
                FROM whatsapp_message_status_events
                WHERE created_at >= ?
            )
            SELECT message_id, recipient_phone, status, template_name, error_title, error_detail, created_at
            FROM latest
            WHERE rn = 1 AND status = 'failed'
            ORDER BY created_at DESC
            LIMIT 8
            """,
            (cutoff_7d,),
        ).fetchall()

        latest_sync_rows = conn.execute(
            """
            SELECT * FROM retargeting_sync_runs
            ORDER BY created_at DESC
            LIMIT 12
            """
        ).fetchall()

        recent_lifecycle_rows = conn.execute(
            """
            SELECT * FROM order_lifecycle_audit
            ORDER BY created_at DESC
            LIMIT 12
            """
        ).fetchall()

        attribution = conn.execute(
            """
            SELECT
                SUM(CASE WHEN COALESCE(meta_lead_id, '') != '' OR COALESCE(fbc, '') != '' OR COALESCE(meta_campaign, '') != '' THEN 1 ELSE 0 END) AS meta_attributed,
                SUM(CASE WHEN COALESCE(google_gclid, '') != '' OR COALESCE(google_campaign, '') != '' THEN 1 ELSE 0 END) AS google_attributed
            FROM customers
            WHERE deleted_at IS NULL
            """
        ).fetchone()

    audience_counts = _sales_audience_counts(audience_rows)
    lifecycle_dict = dict(lifecycle or {})
    shopify_dict = dict(shopify_order_counts or {})
    attribution_dict = dict(attribution or {})
    total_pending_lifecycle = sum(int(lifecycle_dict.get(key) or 0) for key in [
        "order_confirmations_pending",
        "tracking_messages_pending",
        "delivery_messages_pending",
    ])

    action_cards = [
        {
            "key": "order_confirmations_pending",
            "title": "Send Pending Confirmations",
            "count": int(lifecycle_dict.get("order_confirmations_pending") or 0),
            "tone": "red" if int(lifecycle_dict.get("order_confirmations_pending") or 0) else "green",
            "action": "retry_lifecycle:order_created",
        },
        {
            "key": "tracking_messages_pending",
            "title": "Send Tracking Updates",
            "count": int(lifecycle_dict.get("tracking_messages_pending") or 0),
            "tone": "amber" if int(lifecycle_dict.get("tracking_messages_pending") or 0) else "green",
            "action": "retry_lifecycle:tracking_added",
        },
        {
            "key": "delivery_messages_pending",
            "title": "Send Delivered Messages",
            "count": int(lifecycle_dict.get("delivery_messages_pending") or 0),
            "tone": "amber" if int(lifecycle_dict.get("delivery_messages_pending") or 0) else "green",
            "action": "retry_lifecycle:delivered",
        },
        {
            "key": "orders_missing_tracking",
            "title": "Orders Missing Tracking",
            "count": int(lifecycle_dict.get("orders_missing_tracking") or 0),
            "tone": "red" if int(lifecycle_dict.get("orders_missing_tracking") or 0) else "slate",
            "action": "open_journey_orders",
        },
        {
            "key": "message_failures",
            "title": "Failed Messages 7d",
            "count": len(latest_failed),
            "tone": "red" if latest_failed else "green",
            "action": "open_logs_failed",
        },
    ]

    return JSONResponse({
        "ok": True,
        "as_of": now.isoformat(),
        "orders": {
            "shopify_total": int(shopify_dict.get("total_orders") or 0),
            "shopify_today": int(shopify_dict.get("orders_today") or 0),
            "shopify_7d": int(shopify_dict.get("orders_7d") or 0),
            "shopify_fulfilled": int(shopify_dict.get("fulfilled_orders") or 0),
            "shopify_unfulfilled": int(shopify_dict.get("unfulfilled_orders") or 0),
            "journey_total": int(lifecycle_dict.get("total_journey_orders") or 0),
            "automation_orders": int(lifecycle_dict.get("automation_orders") or 0),
            "missing_tracking": int(lifecycle_dict.get("orders_missing_tracking") or 0),
            "in_transit": int(lifecycle_dict.get("in_transit_orders") or 0),
            "delivered": int(lifecycle_dict.get("delivered_orders") or 0),
        },
        "lifecycle": {
            "pending_total": total_pending_lifecycle,
            "order_confirmations_pending": int(lifecycle_dict.get("order_confirmations_pending") or 0),
            "tracking_messages_pending": int(lifecycle_dict.get("tracking_messages_pending") or 0),
            "delivery_messages_pending": int(lifecycle_dict.get("delivery_messages_pending") or 0),
            "recent_audit": [dict(row) for row in recent_lifecycle_rows],
        },
        "automation": {
            "start_at": automation_start,
            "start_label": _automation_start_label(),
            "mode": "today_forward",
            "old_manual_orders_ignored": True,
        },
        "retargeting": {
            "audiences": [
                {"key": key, "label": _audience_label_for_key(key), "count": value}
                for key, value in audience_counts.items()
            ],
            "last_runs": [dict(row) for row in latest_sync_rows],
            "attribution": {
                "meta": int(attribution_dict.get("meta_attributed") or 0),
                "google": int(attribution_dict.get("google_attributed") or 0),
            },
        },
        "failure_queue": [dict(row) for row in latest_failed],
        "action_cards": action_cards,
    })


class _RetargetingSyncRequest(BaseModel):
    audience_key: str = "all"
    platforms: list[str] = Field(default_factory=lambda: ["meta", "google"])
    filters: dict[str, Any] = Field(default_factory=dict)


@router.post("/crm/whatsapp/retargeting/sync")
def sync_whatsapp_retargeting(req: _RetargetingSyncRequest) -> JSONResponse:
    audience_key = (req.audience_key or "all").strip().lower()
    platforms = [p.strip().lower() for p in (req.platforms or []) if p.strip().lower() in {"meta", "google"}]
    if not platforms:
        platforms = ["meta", "google"]

    now_ts = _now_iso()
    with get_db_connection() as conn:
        _ensure_sales_os_tables(conn)
        audience_rows, _summary = _load_whatsapp_audience(conn)
        matched = [row for row in audience_rows if _row_matches_sales_audience(row, audience_key)]

        results: list[dict[str, Any]] = []
        for platform in platforms:
            status = "synced"
            result_payload: dict[str, Any] = {
                "audience_key": audience_key,
                "platform": platform,
                "matched": len(matched),
            }
            if platform == "meta":
                try:
                    from app.services.meta_audience_sync import sync_meta_audiences
                    result_payload["meta_score_tier_sync"] = sync_meta_audiences()
                    if not result_payload["meta_score_tier_sync"].get("ok"):
                        status = "needs_config"
                except Exception as exc:  # noqa: BLE001
                    status = "failed"
                    result_payload["error"] = str(exc)
            else:
                status = "ready_for_google_ads"
                result_payload["note"] = "Audience computed and recorded; Google Ads customer match upload needs Google Ads API credentials."

            run_id = str(uuid.uuid4())
            conn.execute(
                """
                INSERT INTO retargeting_sync_runs
                    (id, audience_key, audience_name, platform, filter_json, matched_count, status, result_json, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    run_id,
                    audience_key,
                    _audience_label_for_key(audience_key),
                    platform,
                    json.dumps(req.filters or {}, ensure_ascii=True),
                    len(matched),
                    status,
                    json.dumps(result_payload, ensure_ascii=True),
                    now_ts,
                    now_ts,
                ),
            )
            results.append({"id": run_id, "platform": platform, "status": status, **result_payload})

    return JSONResponse({
        "ok": all(item["status"] not in {"failed"} for item in results),
        "audience_key": audience_key,
        "audience_name": _audience_label_for_key(audience_key),
        "matched": len(matched),
        "results": results,
    })


class _LifecycleUpdateRequest(BaseModel):
    event_type: str = Field("order_created", description="order_created, tracking_added, delivered")
    shop_domain: str = "rwxtic-gz.myshopify.com"
    order_id: str
    phone: str = ""
    name: str = ""
    email: str = ""
    tracking_id: str = ""
    order_value: float = 0.0
    delivered_at: str | None = None
    send_message: bool = True


def _stage_for_lifecycle_event(event_type: str) -> str:
    event_map = {
        "order_created": "order_confirmed",
        "tracking_added": "in_transit",
        "delivered": "delivered",
    }
    return event_map.get(event_type, "order_confirmed")


@router.post("/crm/whatsapp/orders/lifecycle-update")
def update_order_lifecycle(req: _LifecycleUpdateRequest) -> JSONResponse:
    event_type = (req.event_type or "order_created").strip().lower()
    if event_type not in {"order_created", "tracking_added", "delivered"}:
        return JSONResponse({"ok": False, "error": "Unsupported event_type"}, status_code=400)

    now_ts = _now_iso()
    phone = _normalise_phone(req.phone)
    delivery_status = "pending"
    if event_type == "tracking_added":
        delivery_status = "in_transit"
    elif event_type == "delivered":
        delivery_status = "delivered"

    with get_db_connection() as conn:
        _ensure_sales_os_tables(conn)
        row = conn.execute(
            "SELECT * FROM journey_customers WHERE shop_domain = ? AND shopify_order_id = ? LIMIT 1",
            (req.shop_domain, req.order_id),
        ).fetchone()
        if not row and phone:
            row = conn.execute(
                "SELECT * FROM journey_customers WHERE shop_domain = ? AND phone = ? ORDER BY created_at DESC LIMIT 1",
                (req.shop_domain, phone),
            ).fetchone()

        if row:
            customer_id = row["id"]
            conn.execute(
                """
                UPDATE journey_customers SET
                    phone = COALESCE(NULLIF(?, ''), phone),
                    name = COALESCE(NULLIF(?, ''), name),
                    email = COALESCE(NULLIF(?, ''), email),
                    shopify_order_id = COALESCE(NULLIF(?, ''), shopify_order_id),
                    waybill_id = COALESCE(NULLIF(?, ''), waybill_id),
                    delivery_status = ?,
                    delivered_at = CASE WHEN ? = 'delivered' THEN COALESCE(?, delivered_at, ?) ELSE delivered_at END,
                    order_value_paise = CASE WHEN ? > 0 THEN ? ELSE order_value_paise END,
                    journey_stage = ?,
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    phone,
                    req.name,
                    req.email,
                    req.order_id,
                    req.tracking_id,
                    delivery_status,
                    event_type,
                    req.delivered_at,
                    now_ts,
                    req.order_value,
                    int(req.order_value * 100),
                    _stage_for_lifecycle_event(event_type),
                    now_ts,
                    customer_id,
                ),
            )
        else:
            if not phone:
                _record_lifecycle_audit(
                    conn,
                    event_type=event_type,
                    lifecycle_status="failed",
                    order_id=req.order_id,
                    error_detail="phone is required to create lifecycle order",
                    payload=req.model_dump(),
                )
                return JSONResponse({"ok": False, "error": "phone is required for a new lifecycle order"}, status_code=400)
            customer_id = str(uuid.uuid4())
            conn.execute(
                """
                INSERT INTO journey_customers (
                    id, shop_domain, phone, name, email, shopify_order_id, waybill_id,
                    delivery_status, delivered_at, order_value_paise, journey_stage,
                    journey_started_at, whatsapp_subscription_status, customer_segment,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'subscribed', 'new', ?, ?)
                """,
                (
                    customer_id,
                    req.shop_domain,
                    phone,
                    req.name or None,
                    req.email or None,
                    req.order_id,
                    req.tracking_id or None,
                    delivery_status,
                    req.delivered_at if event_type == "delivered" else None,
                    int(req.order_value * 100),
                    _stage_for_lifecycle_event(event_type),
                    now_ts,
                    now_ts,
                    now_ts,
                ),
            )

        send_payload: dict[str, Any] = {"status": "not_sent", "reason": "send_message=false"}
        if req.send_message:
            try:
                from app.journey_engine import send_journey_message
                fresh_row = conn.execute("SELECT * FROM journey_customers WHERE id = ?", (customer_id,)).fetchone()
                send_result = send_journey_message(conn, dict(fresh_row), _stage_for_lifecycle_event(event_type))
                send_payload = {
                    "status": send_result.status,
                    "template": send_result.template_name,
                    "message_id": send_result.message_id,
                    "error": send_result.error,
                }
                _record_lifecycle_audit(
                    conn,
                    event_type=event_type,
                    lifecycle_status=send_result.status,
                    order_id=req.order_id,
                    customer_id=customer_id,
                    phone=phone,
                    template_name=send_result.template_name,
                    message_id=send_result.message_id,
                    error_detail=send_result.error,
                    payload=req.model_dump(),
                )
            except Exception as exc:  # noqa: BLE001
                send_payload = {"status": "failed", "error": str(exc)}
                _record_lifecycle_audit(
                    conn,
                    event_type=event_type,
                    lifecycle_status="failed",
                    order_id=req.order_id,
                    customer_id=customer_id,
                    phone=phone,
                    error_detail=str(exc),
                    payload=req.model_dump(),
                )

    return JSONResponse({
        "ok": send_payload.get("status") not in {"failed", "error"},
        "customer_id": customer_id,
        "order_id": req.order_id,
        "event_type": event_type,
        "delivery_status": delivery_status,
        "send": send_payload,
    })


class _LifecycleRetryRequest(BaseModel):
    event_type: str = "all"
    limit: int = Field(default=25, ge=1, le=200)
    dry_run: bool = False


@router.post("/crm/whatsapp/orders/retry-lifecycle")
def retry_lifecycle_messages(req: _LifecycleRetryRequest) -> JSONResponse:
    event_type = (req.event_type or "all").strip().lower()
    events = ["order_created", "tracking_added", "delivered"] if event_type == "all" else [event_type]
    if any(event not in {"order_created", "tracking_added", "delivered"} for event in events):
        return JSONResponse({"ok": False, "error": "Unsupported event_type"}, status_code=400)

    results: list[dict[str, Any]] = []
    automation_start = _automation_start_iso()
    with get_db_connection() as conn:
        _ensure_sales_os_tables(conn)
        for event in events:
            if event == "order_created":
                rows = conn.execute(
                    """
                    SELECT * FROM journey_customers
                    WHERE COALESCE(day1_sent, 0) = 0 AND do_not_message = 0 AND whatsapp_subscription_status = 'subscribed'
                                            AND COALESCE(journey_started_at, created_at) >= ?
                    ORDER BY created_at ASC LIMIT ?
                    """,
                                        (automation_start, req.limit),
                ).fetchall()
            elif event == "tracking_added":
                rows = conn.execute(
                    """
                    SELECT * FROM journey_customers
                    WHERE COALESCE(waybill_id, '') != '' AND COALESCE(day2_sent, 0) = 0
                      AND do_not_message = 0 AND whatsapp_subscription_status = 'subscribed'
                                            AND COALESCE(updated_at, created_at) >= ?
                    ORDER BY updated_at ASC LIMIT ?
                    """,
                                        (automation_start, req.limit),
                ).fetchall()
            else:
                rows = conn.execute(
                    """
                    SELECT * FROM journey_customers
                    WHERE delivery_status = 'delivered' AND COALESCE(day5_sent, 0) = 0
                      AND do_not_message = 0 AND whatsapp_subscription_status = 'subscribed'
                                            AND COALESCE(delivered_at, updated_at, created_at) >= ?
                    ORDER BY delivered_at ASC LIMIT ?
                    """,
                                        (automation_start, req.limit),
                ).fetchall()

            sent = failed = dry = 0
            for row in rows:
                customer = dict(row)
                stage = _stage_for_lifecycle_event(event)
                if req.dry_run:
                    dry += 1
                    continue
                try:
                    from app.journey_engine import send_journey_message
                    send_result = send_journey_message(conn, customer, stage)
                    if send_result.status in {"sent", "suppressed"}:
                        sent += 1
                    else:
                        failed += 1
                    _record_lifecycle_audit(
                        conn,
                        event_type=event,
                        lifecycle_status=send_result.status,
                        order_id=customer.get("shopify_order_id"),
                        customer_id=customer.get("id"),
                        phone=customer.get("phone"),
                        template_name=send_result.template_name,
                        message_id=send_result.message_id,
                        error_detail=send_result.error,
                        payload={"retry": True},
                    )
                except Exception as exc:  # noqa: BLE001
                    failed += 1
                    _record_lifecycle_audit(
                        conn,
                        event_type=event,
                        lifecycle_status="failed",
                        order_id=customer.get("shopify_order_id"),
                        customer_id=customer.get("id"),
                        phone=customer.get("phone"),
                        error_detail=str(exc),
                        payload={"retry": True},
                    )
            results.append({"event_type": event, "eligible": len(rows), "sent": sent, "failed": failed, "dry_run": dry, "from_start_at": automation_start})

    return JSONResponse({"ok": True, "automation_start": automation_start, "results": results})


# ─────────────────────────────────────────────────────────────────────────────
# Campaign Planner  —  estimate cost, create campaign, AI suggestions
# ─────────────────────────────────────────────────────────────────────────────

_META_COST_PER_MSG_USD = 0.051  # India marketing rate (Marketing conversation)
_INR_PER_USD = 83.5
_SCORE_BUCKETS: list[tuple[str, int, int]] = [
    ("buyers", 90, 100),
    ("purchase_intent", 80, 89),
    ("warm", 50, 79),
    ("behavioral", 10, 49),
    ("cold", 0, 9),
]


def _normalize_multi_value(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        raw_values = list(value)
    else:
        raw_values = [value]

    normalized: list[str] = []
    for raw_value in raw_values:
        if raw_value is None:
            continue
        for piece in str(raw_value).split(","):
            cleaned = piece.strip()
            if cleaned and cleaned.lower() != "all":
                normalized.append(cleaned)
    return list(dict.fromkeys(normalized))


def _campaign_filters(filters: dict[str, Any] | BaseModel) -> dict[str, Any]:
    payload = filters.model_dump() if isinstance(filters, BaseModel) else dict(filters or {})
    lead_types = _normalize_multi_value(payload.get("lead_types") or payload.get("lead_type"))
    labels = _normalize_multi_value(payload.get("labels"))
    statuses = _normalize_multi_value(payload.get("statuses") or payload.get("status"))
    engagement_labels = _normalize_multi_value(payload.get("engagement_labels") or payload.get("engagement_label"))
    score_buckets = [bucket for bucket in _normalize_multi_value(payload.get("score_buckets")) if bucket in {item[0] for item in _SCORE_BUCKETS}]
    date_field = str(payload.get("date_field") or "updated_at").lower()
    if date_field not in {"created_at", "updated_at", "last_order_date"}:
        date_field = "updated_at"
    delivery_channels = [channel for channel in _normalize_multi_value(payload.get("delivery_channels")) if channel.lower() in {"whatsapp", "email"}]

    return {
        "score_min": max(0, min(100, int(payload.get("score_min") or 0))),
        "score_max": max(0, min(100, int(payload.get("score_max") or 100))),
        "score_buckets": score_buckets,
        "lead_types": [lead_type for lead_type in lead_types if lead_type.lower() != "all"],
        "labels": labels,
        "statuses": [status for status in statuses if status.lower() != "all"],
        "engagement_labels": [label for label in engagement_labels if label.lower() != "all"],
        "language": str(payload.get("language") or "all").lower(),
        "date_field": date_field,
        "date_from": str(payload.get("date_from") or "").strip(),
        "date_to": str(payload.get("date_to") or "").strip(),
        "delivery_channels": delivery_channels or ["whatsapp"],
    }


def _score_bucket_for(score: int) -> str:
    for bucket_id, lower_bound, upper_bound in _SCORE_BUCKETS:
        if lower_bound <= score <= upper_bound:
            return bucket_id
    return "cold"


def _score_bucket_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts = {bucket_id: 0 for bucket_id, _, _ in _SCORE_BUCKETS}
    for row in rows:
        counts[_score_bucket_for(int(row.get("lead_score") or 0))] += 1
    return counts


def _campaign_row_matches(row: dict[str, Any], filters: dict[str, Any]) -> bool:
    score = int(row.get("lead_score") or 0)
    if score < filters["score_min"] or score > filters["score_max"]:
        return False

    score_buckets: list[str] = filters.get("score_buckets") or []
    if score_buckets and _score_bucket_for(score) not in score_buckets:
        return False

    lead_types: list[str] = [str(item).lower() for item in (row.get("lead_types") or []) if item]
    selected_lead_types: list[str] = [str(item).lower() for item in (filters.get("lead_types") or []) if item]
    if selected_lead_types and not set(lead_types).intersection(selected_lead_types):
        return False

    statuses: list[str] = [str(item).lower() for item in (filters.get("statuses") or []) if item]
    if statuses:
        status_tokens = {
            str(row.get("customer_status") or "").lower(),
            str(row.get("purchase_status") or "").lower(),
            str(row.get("engagement_label") or "").lower(),
            *lead_types,
            *[str(tag).lower() for tag in (row.get("tags") or []) if tag],
        }
        if not status_tokens.intersection(statuses):
            return False

    engagement_labels: list[str] = [str(item).lower() for item in (filters.get("engagement_labels") or []) if item]
    if engagement_labels and str(row.get("engagement_label") or "").lower() not in engagement_labels:
        return False

    labels: list[str] = [str(item).lower() for item in (filters.get("labels") or []) if item]
    if labels:
        row_tags = {str(tag).lower() for tag in (row.get("tags") or []) if tag}
        if not row_tags.intersection(labels):
            return False

    date_from = _parse_datetime_like(filters.get("date_from")) if filters.get("date_from") else None
    date_to = _parse_datetime_like(filters.get("date_to")) if filters.get("date_to") else None
    if date_from or date_to:
        row_date = _parse_datetime_like(row.get(filters.get("date_field") or "updated_at"))
        if row_date is None:
            return False
        if date_from and row_date < date_from:
            return False
        if date_to:
            inclusive_to = date_to + timedelta(days=1) if date_to.hour == 0 and date_to.minute == 0 and date_to.second == 0 else date_to
            if row_date > inclusive_to:
                return False

    language = str(filters.get("language") or "all").lower()
    if language == "mal":
        return any(str(tag).lower() == "mal_customer" for tag in (row.get("tags") or []))
    if language == "eng":
        return any(str(tag).lower() == "eng_customer" for tag in (row.get("tags") or []))
    return True


def _filter_campaign_rows(audience_rows: list[dict[str, Any]], filters: dict[str, Any]) -> list[dict[str, Any]]:
    normalized = _campaign_filters(filters)
    return [row for row in audience_rows if _campaign_row_matches(row, normalized)]


class _EstimateCostRequest(BaseModel):
    score_min: int = 0
    score_max: int = 100
    score_buckets: list[str] = Field(default_factory=list)
    lead_types: list[str] = Field(default_factory=list)
    labels: list[str] = Field(default_factory=list)
    statuses: list[str] = Field(default_factory=list)
    engagement_labels: list[str] = Field(default_factory=list)
    lead_type: str = "all"
    language: str = "all"
    date_field: str = "updated_at"
    date_from: str = ""
    date_to: str = ""
    delivery_channels: list[str] = Field(default_factory=lambda: ["whatsapp"])


@router.get("/crm/whatsapp/labels")
def whatsapp_campaign_labels(limit: int = Query(default=100, ge=1, le=500), search: str = Query(default="")) -> JSONResponse:
    with get_db_connection() as conn:
        audience_rows, _summary = _load_whatsapp_audience(conn)

    counter: Counter[str] = Counter()
    for row in audience_rows:
        for tag in row.get("tags") or []:
            tag_value = str(tag).strip()
            if tag_value:
                counter[tag_value] += 1

    needle = search.strip().lower()
    labels = [
        {"label": label, "count": count}
        for label, count in counter.most_common()
        if not needle or needle in label.lower()
    ]

    return JSONResponse({"ok": True, "labels": labels[:limit]})


@router.post("/crm/whatsapp/campaigns/estimate-cost")
@router.post("/crm/campaigns/estimate-cost")
def estimate_campaign_cost(req: _EstimateCostRequest) -> JSONResponse:
    with get_db_connection() as conn:
        audience_rows, _summary = _load_whatsapp_audience(conn)

    filters = _campaign_filters(req)
    filtered = _filter_campaign_rows(audience_rows, filters)

    whatsapp_sendable = [r for r in filtered if r.get("whatsapp_opted_in") or r.get("phone")]
    email_sendable = [r for r in filtered if _display_email(r.get("email"))]
    send_count = len(whatsapp_sendable)

    score_breakdown = _score_bucket_counts(filtered)
    cost_usd = round(send_count * _META_COST_PER_MSG_USD, 2)
    cost_inr = round(cost_usd * _INR_PER_USD, 0)

    return JSONResponse({
        "ok": True,
        "matched": len(filtered),
        "opted_in": send_count,
        "whatsapp_sendable": len(whatsapp_sendable),
        "email_sendable": len(email_sendable),
        "delivery_channels": filters.get("delivery_channels") or ["whatsapp"],
        "language_split": {
            "mal": sum(1 for r in filtered if any(str(t).lower() == "mal_customer" for t in (r.get("tags") or []))),
            "eng": sum(1 for r in filtered if any(str(t).lower() == "eng_customer" for t in (r.get("tags") or []))),
        },
        "score_breakdown": score_breakdown,
        "cost": {
            "usd": cost_usd,
            "inr": int(cost_inr),
            "per_message_usd": _META_COST_PER_MSG_USD,
        },
        "filters": filters,
    })


class _CreateCampaignRequest(BaseModel):
    name: str
    audience_filters: dict
    template_id: str
    template_name: str
    channel: str = "wabis"
    delivery_channels: list[str] = Field(default_factory=lambda: ["whatsapp"])
    email_subject: str = ""
    email_html: str = ""
    schedule_type: str = "now"
    scheduled_at: str | None = None
    sync_meta_audiences: bool = False


def _default_crm_email_html(campaign_name: str, subject: str) -> str:
    return (
        "<div style='font-family:Arial,sans-serif;max-width:620px;margin:0 auto;padding:24px;color:#1f2937;'>"
        f"<h1 style='font-size:24px;margin:0 0 12px;color:#0f766e;'>{campaign_name}</h1>"
        f"<p style='font-size:16px;line-height:1.7;margin:0 0 18px;'>{subject}</p>"
        "<p style='font-size:15px;line-height:1.7;margin:0 0 20px;'>"
        "Explore fresh, authentic Pureleven products selected for your kitchen and wellness routine."
        "</p>"
        "<a href='https://pureleven.com/collections/all?utm_source=email&utm_medium=crm&utm_campaign=custom_audience' "
        "style='display:inline-block;background:#0f766e;color:#ffffff;text-decoration:none;padding:12px 18px;border-radius:8px;font-weight:700;'>Shop Pureleven</a>"
        "</div>"
    )


def _queue_email_campaign_for_audience(
    *,
    name: str,
    subject: str,
    html_body: str,
    recipients: list[dict[str, Any]],
    scheduled_at: str,
    schedule_type: str,
    filters: dict[str, Any],
) -> dict[str, Any]:
    campaign_id = f"crm_email_{uuid.uuid4().hex[:12]}"
    now_ts = _now_iso()
    subject_text = subject.strip() or f"Pureleven update: {name}"
    html = html_body.strip() or _default_crm_email_html(name, subject_text)
    status = "scheduled" if schedule_type == "scheduled" else "queued"
    queued = 0
    seen: set[str] = set()
    with get_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO promotional_campaigns
                (campaign_id, name, template_type, subject, html_body, discount_pct, coupon_code,
                 segment, status, queue_status, queued_count, send_interval_seconds,
                 scheduled_at, created_at, updated_at)
            VALUES (?, ?, 'crm_custom_audience', ?, ?, 0, '', ?, ?, ?, 0, 1.0, ?, ?, ?)
            """,
            (
                campaign_id,
                name,
                subject_text,
                html,
                json.dumps(filters, ensure_ascii=True),
                status,
                status,
                scheduled_at,
                now_ts,
                now_ts,
            ),
        )
        for row in recipients:
            email = (_display_email(row.get("email")) or "").strip().lower()
            if not email or email in seen:
                continue
            seen.add(email)
            conn.execute(
                """
                INSERT OR IGNORE INTO campaign_send_queue
                    (queue_id, campaign_id, email, first_name, last_name, status, scheduled_for,
                     attempt_count, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, 'queued', ?, 0, ?, ?)
                """,
                (
                    str(uuid.uuid4()),
                    campaign_id,
                    email,
                    row.get("first_name") or row.get("full_name") or "",
                    row.get("last_name") or "",
                    scheduled_at,
                    now_ts,
                    now_ts,
                ),
            )
            queued += 1
        conn.execute(
            "UPDATE promotional_campaigns SET queued_count = ?, updated_at = ? WHERE campaign_id = ?",
            (queued, _now_iso(), campaign_id),
        )
    return {"campaign_id": campaign_id, "queued": queued, "status": status, "subject": subject_text}


@router.post("/crm/whatsapp/journey/create-campaign")
@router.post("/crm/journey/create-campaign")
def create_journey_campaign(req: _CreateCampaignRequest) -> JSONResponse:
    with get_db_connection() as conn:
        audience_rows, _summary = _load_whatsapp_audience(conn)

    filters_payload = dict(req.audience_filters or {})
    filters_payload["delivery_channels"] = req.delivery_channels or filters_payload.get("delivery_channels") or ["whatsapp"]
    filters = _campaign_filters(filters_payload)
    filtered = _filter_campaign_rows(audience_rows, filters)

    delivery_channels = filters.get("delivery_channels") or ["whatsapp"]
    whatsapp_sendable = [r for r in filtered if r.get("phone")]
    email_sendable = [r for r in filtered if _display_email(r.get("email"))]
    sendable = whatsapp_sendable if "whatsapp" in delivery_channels else []
    campaign_id = str(uuid.uuid4())
    now_ts = datetime.now(timezone.utc).isoformat()
    scheduled_at = req.scheduled_at or now_ts
    email_campaign: dict[str, Any] | None = None

    with get_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO campaigns (
                id, name, description, type, audience_type, segment_filter_json,
                template_id, schedule_type, scheduled_at, status,
                total_recipients, sent_count, error_count, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, ?, ?)
            """,
            (
                campaign_id,
                req.name,
                f"CRM audience campaign via {', '.join(delivery_channels)} — {req.template_name or 'email only'}",
                "omnichannel" if set(delivery_channels) == {"whatsapp", "email"} else delivery_channels[0],
                req.audience_filters.get("lead_type", "all"),
                json.dumps(filters, ensure_ascii=True),
                req.template_id,
                req.schedule_type,
                scheduled_at,
                "pending" if req.schedule_type == "now" else "scheduled",
                len(sendable) + (len(email_sendable) if "email" in delivery_channels else 0),
                now_ts,
                now_ts,
            ),
        )
        for r in sendable:
            phone = r.get("phone") or ""
            conn.execute(
                """
                INSERT OR IGNORE INTO campaign_send_queue (
                    queue_id, campaign_id, email, first_name, status, scheduled_for, attempt_count, created_at, updated_at
                ) VALUES (?, ?, ?, ?, 'pending', ?, 0, ?, ?)
                """,
                (
                    str(uuid.uuid4()),
                    campaign_id,
                    f"whatsapp:{phone}",
                    r.get("first_name") or r.get("name") or "",
                    scheduled_at,
                    now_ts,
                    now_ts,
                ),
            )
        conn.commit()

    if "email" in delivery_channels:
        email_campaign = _queue_email_campaign_for_audience(
            name=req.name,
            subject=req.email_subject,
            html_body=req.email_html,
            recipients=email_sendable,
            scheduled_at=scheduled_at,
            schedule_type=req.schedule_type,
            filters=filters,
        )

    if req.sync_meta_audiences:
        try:
            from app.services.meta_audience_sync import sync_meta_audiences
            sync_meta_audiences()
        except Exception:  # noqa: BLE001
            pass

    return JSONResponse({
        "ok": True,
        "campaign_id": campaign_id,
        "email_campaign_id": email_campaign.get("campaign_id") if email_campaign else None,
        "name": req.name,
        "recipients": len(sendable) + (email_campaign.get("queued", 0) if email_campaign else 0),
        "whatsapp_recipients": len(sendable),
        "email_recipients": email_campaign.get("queued", 0) if email_campaign else 0,
        "delivery_channels": delivery_channels,
        "matched": len(filtered),
        "filters": filters,
        "schedule_type": req.schedule_type,
        "scheduled_at": scheduled_at,
        "status": "pending" if req.schedule_type == "now" else "scheduled",
    })


@router.get("/crm/whatsapp/journey/campaigns")
@router.get("/crm/journey/campaigns")
def list_journey_campaigns(
    limit: int = Query(default=12, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
) -> JSONResponse:
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            WITH queue_rollup AS (
                SELECT
                    campaign_id,
                    COUNT(*) AS queue_total,
                    SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) AS pending_count,
                    SUM(CASE WHEN status = 'queued' THEN 1 ELSE 0 END) AS queued_count,
                    SUM(CASE WHEN status = 'sending' THEN 1 ELSE 0 END) AS sending_count,
                    SUM(CASE WHEN status = 'sent' THEN 1 ELSE 0 END) AS queue_sent_count,
                    SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) AS queue_failed_count,
                    SUM(CASE WHEN status = 'suppressed' THEN 1 ELSE 0 END) AS suppressed_count,
                    MAX(COALESCE(sent_at, updated_at, created_at)) AS last_queue_event_at
                FROM campaign_send_queue
                GROUP BY campaign_id
            ),
            send_rollup AS (
                SELECT
                    campaign_id,
                    COUNT(*) AS send_total,
                    SUM(CASE WHEN status = 'sent' THEN 1 ELSE 0 END) AS send_sent_count,
                    SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) AS send_failed_count,
                    SUM(CASE WHEN opened_at IS NOT NULL THEN 1 ELSE 0 END) AS opened_count,
                    SUM(CASE WHEN clicked_at IS NOT NULL THEN 1 ELSE 0 END) AS clicked_count,
                    MAX(COALESCE(sent_at, created_at)) AS last_send_event_at
                FROM campaign_sends
                GROUP BY campaign_id
            )
            SELECT
                c.*,
                qr.queue_total,
                qr.pending_count,
                qr.queued_count,
                qr.sending_count,
                qr.queue_sent_count,
                qr.queue_failed_count,
                qr.suppressed_count,
                qr.last_queue_event_at,
                sr.send_total,
                sr.send_sent_count,
                sr.send_failed_count,
                sr.opened_count,
                sr.clicked_count,
                sr.last_send_event_at
            FROM campaigns c
            LEFT JOIN queue_rollup qr ON qr.campaign_id = c.id
            LEFT JOIN send_rollup sr ON sr.campaign_id = c.id
            WHERE c.type IN ('whatsapp', 'omnichannel', 'email')
            ORDER BY COALESCE(c.updated_at, c.created_at) DESC
            LIMIT ? OFFSET ?
            """,
            (limit, offset),
        ).fetchall()
        total = conn.execute(
            "SELECT COUNT(*) AS n FROM campaigns WHERE type IN ('whatsapp', 'omnichannel', 'email')"
        ).fetchone()["n"]

    items: list[dict[str, Any]] = []
    for row in rows:
        data = dict(row)
        filters = _json_loads(data.get("segment_filter_json"), {})
        items.append({
            "id": data.get("id"),
            "name": data.get("name"),
            "description": data.get("description"),
            "type": data.get("type"),
            "status": data.get("status"),
            "template_id": data.get("template_id"),
            "schedule_type": data.get("schedule_type"),
            "scheduled_at": data.get("scheduled_at"),
            "started_at": data.get("started_at"),
            "completed_at": data.get("completed_at"),
            "created_at": data.get("created_at"),
            "updated_at": data.get("updated_at"),
            "total_recipients": data.get("total_recipients") or 0,
            "sent_count": data.get("sent_count") or 0,
            "error_count": data.get("error_count") or 0,
            "filters": filters,
            "queue": {
                "total": data.get("queue_total") or 0,
                "pending": data.get("pending_count") or 0,
                "queued": data.get("queued_count") or 0,
                "sending": data.get("sending_count") or 0,
                "sent": data.get("queue_sent_count") or 0,
                "failed": data.get("queue_failed_count") or 0,
                "suppressed": data.get("suppressed_count") or 0,
            },
            "sends": {
                "total": data.get("send_total") or 0,
                "sent": data.get("send_sent_count") or 0,
                "failed": data.get("send_failed_count") or 0,
                "opened": data.get("opened_count") or 0,
                "clicked": data.get("clicked_count") or 0,
            },
            "last_activity_at": _pick_latest_text(
                data.get("completed_at"),
                data.get("last_send_event_at"),
                data.get("last_queue_event_at"),
                data.get("updated_at"),
                data.get("created_at"),
            ),
        })

    return JSONResponse({
        "ok": True,
        "items": items,
        "total": total,
        "count": len(items),
        "limit": limit,
        "offset": offset,
    })


@router.post("/crm/whatsapp/ai/suggest-campaign")
@router.post("/crm/ai/suggest-campaign")
def ai_suggest_campaign(req: _EstimateCostRequest) -> JSONResponse:
    """Return top high-value customers ranked by a composite hot-score."""
    now = datetime.now(timezone.utc)

    with get_db_connection() as conn:
        customer_rows = _safe_fetchall(
            conn,
            """
            SELECT
                id, email, phone, name, first_name, whatsapp_number, tags,
                lead_score, purchase_status, last_order_date, last_engagement_at,
                total_orders, total_spent, engagement_label
            FROM customers
            WHERE deleted_at IS NULL
              AND COALESCE(whatsapp_number, phone, '') != ''
              AND COALESCE(lead_score, 0) >= 40
            ORDER BY COALESCE(lead_score, 0) DESC
            LIMIT 2000
            """,
        )
        wabis_templates = _get_cached_templates("wabis")
        meta_templates = _get_cached_templates("meta")

    def _days_ago(ts: str | None) -> float:
        if not ts:
            return 365.0
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return max(0.0, (now - dt).total_seconds() / 86400)
        except Exception:  # noqa: BLE001
            return 365.0

    filters = _campaign_filters(req)
    filtered_rows = _filter_campaign_rows([dict(row) for row in customer_rows], filters)

    template_candidates = [*wabis_templates, *meta_templates]

    def _template_score(template: dict[str, Any]) -> int:
        text_blob = " ".join(
            str(part or "")
            for part in [
                template.get("name"),
                template.get("category"),
                template.get("header", {}).get("text") if isinstance(template.get("header"), dict) else None,
                template.get("body", {}).get("meta_text") if isinstance(template.get("body"), dict) else None,
                template.get("body", {}).get("text") if isinstance(template.get("body"), dict) else None,
            ]
        ).lower()
        score = 0
        if any(bucket in ["buyers", "purchase_intent"] for bucket in filters["score_buckets"]):
            for keyword in ["offer", "discount", "deal", "reorder", "buy", "premium", "exclusive", "sale"]:
                if keyword in text_blob:
                    score += 4
        if any(bucket in ["warm", "behavioral"] for bucket in filters["score_buckets"]):
            for keyword in ["follow", "reminder", "nudge", "help", "welcome", "check", "update"]:
                if keyword in text_blob:
                    score += 3
        if any(bucket == "cold" for bucket in filters["score_buckets"]):
            for keyword in ["welcome", "intro", "hello", "discover", "browse"]:
                if keyword in text_blob:
                    score += 3
        if filters["lead_types"] and "purchased" in filters["lead_types"]:
            if "order" in text_blob or "reorder" in text_blob or "purchase" in text_blob:
                score += 5
        if filters["labels"] and any(label in text_blob for label in filters["labels"]):
            score += 2
        if template.get("send_issue"):
            score -= 10
        return score

    recommended_template = None
    if template_candidates:
        recommended_template = max(template_candidates, key=_template_score)

    if not filtered_rows:
        filtered_rows = [dict(row) for row in customer_rows[:50]]

    suggestions = []
    for row in filtered_rows:
        d = dict(row)
        score = int(d.get("lead_score") or 0)
        order_recency = max(0.0, 1.0 - _days_ago(d.get("last_order_date")) / 90)
        engagement_recency = max(0.0, 1.0 - _days_ago(d.get("last_engagement_at")) / 30)
        hot_score = round(score * 0.5 + engagement_recency * 30 + order_recency * 20, 1)
        tags = _split_tags(d.get("tags"))
        language = "mal" if any(str(t).lower() == "mal_customer" for t in tags) else "eng"
        suggestions.append({
            "id": d.get("id"),
            "name": d.get("name") or d.get("first_name") or "—",
            "phone": d.get("whatsapp_number") or d.get("phone"),
            "email": d.get("email"),
            "lead_score": score,
            "hot_score": hot_score,
            "purchase_status": d.get("purchase_status"),
            "engagement_label": d.get("engagement_label"),
            "total_orders": d.get("total_orders") or 0,
            "language": language,
        })

    suggestions.sort(key=lambda x: x["hot_score"], reverse=True)
    top = suggestions[:20]

    buyer_count = sum(1 for s in top if s.get("purchase_status") == "purchased")
    mal_count = sum(1 for s in top if s.get("language") == "mal")
    recommended_language = "mal" if mal_count > len(top) / 2 else "eng"
    recommendation = "Upsell recent buyers" if any(bucket in ["buyers", "purchase_intent"] for bucket in filters["score_buckets"]) else "Re-engage warm and behavioral leads"

    return JSONResponse({
        "ok": True,
        "recommendation": recommendation,
        "filters": filters,
        "top_customers": top,
        "summary": {
            "total_hot": len(suggestions),
            "top_20_buyers": buyer_count,
            "recommended_language": recommended_language,
            "recommended_score_min": 70,
            "recommended_score_max": 100,
            "estimated_conversion_pct": round(buyer_count / max(len(top), 1) * 100, 0),
        },
        "recommended_template": (
            {
                "id": recommended_template.get("id"),
                "name": recommended_template.get("name"),
                "channel": "meta" if recommended_template in meta_templates else "wabis",
                "category": recommended_template.get("category"),
                "locale": recommended_template.get("locale"),
                "body": recommended_template.get("body"),
                "header": recommended_template.get("header"),
                "footer": recommended_template.get("footer"),
            }
            if recommended_template
            else None
        ),
    })


@router.get("/crm/whatsapp/journey/customers")
def whatsapp_journey_customers(
    search: str = Query(default=""),
    phone: str = Query(default=""),
    stage: str = Query(default="all"),
    segment: str = Query(default="all"),
    subscription: str = Query(default="all"),
    limit: int = Query(default=80, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> JSONResponse:
    scope_sql, scope_params = _journey_scope("jc")
    with get_db_connection() as conn:
        _bootstrap_journey_customers_from_audience(conn)

    where = [scope_sql]
    params: list[Any] = [*scope_params]

    if stage != "all":
        where.append("jc.journey_stage = ?")
        params.append(stage)
    if segment != "all":
        where.append("jc.customer_segment = ?")
        params.append(segment)
    if subscription != "all":
        where.append("jc.whatsapp_subscription_status = ?")
        params.append(subscription)
    phone_order_needle = ""
    if search.strip():
        needle = f"%{search.strip().lower()}%"
        where.append(
            "(" 
            "LOWER(COALESCE(jc.name, '')) LIKE ? OR "
            "LOWER(COALESCE(jc.email, '')) LIKE ? OR "
            "REPLACE(COALESCE(jc.phone, ''), '+', '') LIKE ? OR "
            "LOWER(COALESCE(jc.shopify_order_id, '')) LIKE ?"
            ")"
        )
        phone_needle = re.sub(r"\D", "", search)
        params.extend([needle, needle, f"%{phone_needle}%" if phone_needle else "%", needle])
    if phone.strip():
        phone_order_needle = re.sub(r"\D", "", phone)
        if phone_order_needle:
            normalized_phone_sql = (
                "REPLACE(REPLACE(REPLACE(COALESCE(jc.phone, ''), '+', ''), ' ', ''), '-', '')"
            )
            where.append(f"{normalized_phone_sql} LIKE ?")
            params.append(f"%{phone_order_needle}%")

    where_sql = " AND ".join(where)
    base_sql = f"""
        WITH message_stats AS (
            SELECT customer_id, COUNT(*) AS total_messages, MAX(sent_at) AS last_sent_at
            FROM journey_messages
            GROUP BY customer_id
        ),
        engagement_stats AS (
            SELECT customer_id, COUNT(*) AS total_events, SUM(points_awarded) AS total_points, MAX(created_at) AS last_event_at
            FROM journey_engagement_events
            GROUP BY customer_id
        ),
        status_stats AS (
            SELECT jm.customer_id, MAX(wm.created_at) AS last_status_at,
                   SUM(CASE WHEN wm.status = 'failed' THEN 1 ELSE 0 END) AS failed_count
            FROM journey_messages jm
            JOIN whatsapp_message_status_events wm ON wm.message_id = jm.wabis_message_id
            GROUP BY jm.customer_id
        )
        SELECT
            jc.id,
            jc.phone,
            jc.name,
            jc.email,
            jc.journey_stage,
            jc.customer_segment,
            jc.whatsapp_subscription_status,
            jc.delivery_status,
            jc.google_review_status,
            jc.engagement_score,
            jc.purchase_count,
            ROUND(COALESCE(jc.total_spent_paise, 0) / 100.0, 2) AS total_spent,
            jc.last_purchase_at,
            jc.last_engagement_at,
            jc.updated_at,
            COALESCE(ms.total_messages, 0) AS total_messages,
            ms.last_sent_at,
            COALESCE(es.total_events, 0) AS total_events,
            ROUND(COALESCE(es.total_points, 0), 1) AS total_points,
            ss.last_status_at,
            COALESCE(ss.failed_count, 0) AS failed_count
        FROM journey_customers jc
        LEFT JOIN message_stats ms ON ms.customer_id = jc.id
        LEFT JOIN engagement_stats es ON es.customer_id = jc.id
        LEFT JOIN status_stats ss ON ss.customer_id = jc.id
        WHERE {where_sql}
    """

    with get_db_connection() as conn:
        order_sql = "ORDER BY jc.updated_at DESC"
        if phone_order_needle:
            order_sql = (
                "ORDER BY CASE WHEN REPLACE(REPLACE(REPLACE(COALESCE(jc.phone, ''), '+', ''), ' ', ''), '-', '') LIKE ? THEN 0 ELSE 1 END, jc.updated_at DESC"
            )
        rows = conn.execute(
            base_sql + f" {order_sql} LIMIT ? OFFSET ?",
            [*params, *([f"%{phone_order_needle}%"] if phone_order_needle else []), limit, offset],
        ).fetchall()
        total = conn.execute(
            f"SELECT COUNT(*) AS n FROM ({base_sql})",
            params,
        ).fetchone()["n"]

    return JSONResponse({
        "ok": True,
        "customers": [dict(row) for row in rows],
        "total": total,
        "count": len(rows),
    })


@router.get("/crm/whatsapp/journey/{customer_id}")
def whatsapp_journey_detail(customer_id: str) -> JSONResponse:
    scope_sql, scope_params = _journey_scope("jc")
    with get_db_connection() as conn:
        _bootstrap_journey_customers_from_audience(conn)
        customer = conn.execute(
            f"SELECT * FROM journey_customers jc WHERE jc.id = ? AND {scope_sql} LIMIT 1",
            [customer_id, *scope_params],
        ).fetchone()
        if not customer:
            return JSONResponse({"ok": False, "detail": "Customer not found"}, status_code=404)

        message_rows = conn.execute(
            """
            WITH latest_status AS (
                SELECT
                    message_id,
                    status,
                    error_title,
                    error_detail,
                    created_at,
                    ROW_NUMBER() OVER (PARTITION BY message_id ORDER BY created_at DESC) AS rn
                FROM whatsapp_message_status_events
            )
            SELECT
                jm.sent_at AS event_at,
                'template_sent' AS event_type,
                jm.template_name,
                jm.journey_stage,
                COALESCE(ls.status, jm.delivery_status, 'sent') AS status,
                ls.error_title,
                ls.error_detail,
                jm.variables_json
            FROM journey_messages jm
            LEFT JOIN latest_status ls ON ls.message_id = jm.wabis_message_id AND ls.rn = 1
            WHERE jm.customer_id = ?
            ORDER BY jm.sent_at DESC
            LIMIT 200
            """,
            (customer_id,),
        ).fetchall()

        engagement_rows = conn.execute(
            """
            SELECT
                created_at AS event_at,
                event_type,
                template_name,
                journey_stage,
                points_awarded,
                metadata_json
            FROM journey_engagement_events
            WHERE customer_id = ?
            ORDER BY created_at DESC
            LIMIT 200
            """,
            (customer_id,),
        ).fetchall()

        conversation_rows = conn.execute(
            """
            SELECT
                cm.created_at AS event_at,
                CASE WHEN cm.actor = 'customer' THEN 'customer_reply' ELSE 'assistant_message' END AS event_type,
                NULL AS template_name,
                NULL AS journey_stage,
                cm.message_type,
                cm.customer_text,
                cm.message_rendered,
                cm.delivery_status
            FROM conversation_sessions cs
            JOIN conversation_messages cm ON cm.session_id = cs.id
            WHERE cs.customer_id = ?
            ORDER BY cm.created_at DESC
            LIMIT 200
            """,
            (customer_id,),
        ).fetchall()

    timeline: list[dict[str, Any]] = []
    for row in message_rows:
        timeline.append({
            "source": "message",
            "event_at": row["event_at"],
            "event_type": row["event_type"],
            "template_name": row["template_name"],
            "journey_stage": row["journey_stage"],
            "status": row["status"],
            "error_title": row["error_title"],
            "error_detail": row["error_detail"],
            "variables": _json_loads(row["variables_json"], {}),
        })
    for row in engagement_rows:
        timeline.append({
            "source": "engagement",
            "event_at": row["event_at"],
            "event_type": row["event_type"],
            "template_name": row["template_name"],
            "journey_stage": row["journey_stage"],
            "points_awarded": row["points_awarded"],
            "metadata": _json_loads(row["metadata_json"], {}),
        })
    for row in conversation_rows:
        timeline.append({
            "source": "conversation",
            "event_at": row["event_at"],
            "event_type": row["event_type"],
            "message_type": row["message_type"],
            "customer_text": row["customer_text"],
            "message_rendered": row["message_rendered"],
            "status": row["delivery_status"],
        })

    timeline.sort(key=lambda item: item.get("event_at") or "", reverse=True)

    customer_dict = dict(customer)
    customer_dict["total_spent"] = round(float(customer_dict.get("total_spent_paise") or 0) / 100.0, 2)
    return JSONResponse({
        "ok": True,
        "customer": customer_dict,
        "timeline": timeline,
        "timeline_count": len(timeline),
    })


@router.get("/crm/whatsapp/logs")
def whatsapp_logs(
    search: str = Query(default=""),
    status: str = Query(default="all"),
    template_name: str = Query(default=""),
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> JSONResponse:
    params: list[Any] = []
    outer_where: list[str] = ["1=1"]
    if status != "all":
        outer_where.append("status = ?")
        params.append(status)
    if template_name.strip():
        outer_where.append("LOWER(COALESCE(template_name, '')) LIKE ?")
        params.append(f"%{template_name.strip().lower()}%")
    if search.strip():
        needle = f"%{search.strip().lower()}%"
        phone_needle = re.sub(r"\D", "", search)
        outer_where.append(
            "(" 
            "LOWER(COALESCE(customer_name, '')) LIKE ? OR "
            "LOWER(COALESCE(customer_email, '')) LIKE ? OR "
            "LOWER(COALESCE(template_name, '')) LIKE ? OR "
            "REPLACE(COALESCE(recipient_phone, ''), '+', '') LIKE ?"
            ")"
        )
        params.extend([needle, needle, needle, f"%{phone_needle}%" if phone_needle else "%"])

    base_sql = """
        WITH latest_status AS (
            SELECT
                message_id,
                recipient_phone,
                status,
                template_name,
                error_code,
                error_title,
                error_detail,
                created_at,
                ROW_NUMBER() OVER (PARTITION BY message_id ORDER BY created_at DESC) AS rn
            FROM whatsapp_message_status_events
        ),
        journey_log AS (
            SELECT
                COALESCE(ls.message_id, jm.wabis_message_id, jm.id) AS log_id,
                COALESCE(ls.created_at, jm.sent_at) AS event_at,
                COALESCE(ls.status, jm.delivery_status, 'sent') AS status,
                COALESCE(ls.template_name, jm.template_name) AS template_name,
                jc.phone AS recipient_phone,
                jc.name AS customer_name,
                jc.email AS customer_email,
                jm.journey_stage,
                ls.error_code,
                ls.error_title,
                ls.error_detail,
                'journey' AS source
            FROM journey_messages jm
            JOIN journey_customers jc ON jc.id = jm.customer_id
            LEFT JOIN latest_status ls ON ls.message_id = jm.wabis_message_id AND ls.rn = 1
        ),
        manual_log AS (
            SELECT
                ls.message_id AS log_id,
                ls.created_at AS event_at,
                ls.status,
                ls.template_name,
                ls.recipient_phone,
                NULL AS customer_name,
                NULL AS customer_email,
                NULL AS journey_stage,
                ls.error_code,
                ls.error_title,
                ls.error_detail,
                'manual' AS source
            FROM latest_status ls
            WHERE ls.rn = 1 AND NOT EXISTS (
                SELECT 1 FROM journey_messages jm WHERE jm.wabis_message_id = ls.message_id
            )
        )
        SELECT *
        FROM (
            SELECT * FROM journey_log
            UNION ALL
            SELECT * FROM manual_log
        )
    """

    where_sql = " AND ".join(outer_where)
    with get_db_connection() as conn:
        rows = conn.execute(
            f"{base_sql} WHERE {where_sql} ORDER BY event_at DESC LIMIT ? OFFSET ?",
            [*params, limit, offset],
        ).fetchall()
        total = conn.execute(
            f"SELECT COUNT(*) AS n FROM ({base_sql} WHERE {where_sql})",
            params,
        ).fetchone()["n"]
        summary_rows = conn.execute(
            f"SELECT status, COUNT(*) AS cnt FROM ({base_sql} WHERE {where_sql}) GROUP BY status ORDER BY cnt DESC",
            params,
        ).fetchall()

    return JSONResponse({
        "ok": True,
        "logs": [dict(row) for row in rows],
        "total": total,
        "count": len(rows),
        "summary": {row["status"]: row["cnt"] for row in summary_rows},
    })


@router.post("/crm/whatsapp/customers/import/shopify")
def import_whatsapp_customers_shopify() -> JSONResponse:
    try:
        importer = ShopifyCustomerImporter()
        customer_rows = importer.get_all_customers()
        order_rows = [row for row in importer.get_order_contacts(days=3650, limit_pages=50) if (row.get("phone") or "").strip()]
    except Exception as exc:
        logger.error("WhatsApp Shopify import fetch failed: %s", exc)
        return JSONResponse({"ok": False, "detail": str(exc)}, status_code=503)

    customers = _merge_shopify_contacts(customer_rows, order_rows)
    result = _import_contacts(customers, source="shopify_whatsapp_import", default_tags=["whatsapp", "shopify"])
    result["source"] = "shopify"
    result["shopify_customer_rows"] = len(customer_rows)
    result["shopify_order_contact_rows"] = len(order_rows)
    if result["total_rows"] > 0 and result["skipped_no_phone"] == result["total_rows"]:
        result["detail"] = "Shopify returned customer and order records, but none contained a phone number usable for WhatsApp import."
    return JSONResponse(result)


@router.post("/crm/whatsapp/customers/import/csv")
async def import_whatsapp_customers_csv(
    file: UploadFile = File(...),
    default_tags: str = Query(default=""),
) -> JSONResponse:
    raw = await file.read()
    rows = _read_csv_rows(raw)
    result = _import_contacts(rows, source="whatsapp_csv_import", default_tags=_split_tags(default_tags) or ["whatsapp", "csv"])
    result["source"] = "csv"
    return JSONResponse(result)


@router.post("/crm/whatsapp/customers/import/xml")
async def import_whatsapp_customers_xml(
    file: UploadFile = File(...),
    default_tags: str = Query(default=""),
) -> JSONResponse:
    raw = await file.read()
    try:
        rows = _read_xml_rows(raw)
    except Exception as exc:
        return JSONResponse({"ok": False, "detail": f"Invalid XML file: {exc}"}, status_code=400)
    result = _import_contacts(rows, source="whatsapp_xml_import", default_tags=_split_tags(default_tags) or ["whatsapp", "xml"])
    result["source"] = "xml"
    return JSONResponse(result)


@router.post("/crm/whatsapp/customers/import/xlsx")
async def import_whatsapp_customers_xlsx(
    file: UploadFile = File(...),
    default_tags: str = Query(default=""),
) -> JSONResponse:
    raw = await file.read()
    try:
        rows = _read_xlsx_rows(raw)
    except RuntimeError as exc:
        return JSONResponse({"ok": False, "detail": str(exc)}, status_code=503)
    except Exception as exc:
        return JSONResponse({"ok": False, "detail": f"Invalid XLSX file: {exc}"}, status_code=400)
    result = _import_contacts(rows, source="whatsapp_xlsx_import", default_tags=_split_tags(default_tags) or ["whatsapp", "xlsx"])
    result["source"] = "xlsx"
    return JSONResponse(result)


@router.post("/crm/meta/audiences/sync")
def trigger_meta_audience_sync() -> JSONResponse:
    """Manually trigger Meta Custom Audience sync for all score tiers."""
    try:
        from app.services.meta_audience_sync import sync_meta_audiences
        result = sync_meta_audiences()
        return JSONResponse(result)
    except Exception as exc:
        logger.error("meta_audience_sync manual trigger failed: %s", exc)
        return JSONResponse({"ok": False, "detail": str(exc)}, status_code=503)


@router.post("/crm/scores/recompute")
def trigger_score_recompute() -> JSONResponse:
    """Manually trigger a full customer score recomputation (all customers)."""
    try:
        from app.services.customer_intelligence_service import recompute_all_customer_scores
        result = recompute_all_customer_scores(limit=25000)
        return JSONResponse({"ok": True, **result})
    except Exception as exc:
        logger.error("score recompute manual trigger failed: %s", exc)
        return JSONResponse({"ok": False, "detail": str(exc)}, status_code=503)


# ─── Send Models ──────────────────────────────────────────────────────────────

class WabisSendRequest(BaseModel):
    phone: str
    template_id: int | None = None
    template_name: str = ""
    params: list[str] = []
    header_image_url: str = ""


class MetaSendRequest(BaseModel):
    phone: str
    template: str
    params: list[str] = []
    phone_number_id: str = ""
    language_code: str = "en_US"


# ─── Send via Wabis ──────────────────────────────────────────────────────────

@router.post("/crm/wabis/send")
def send_via_wabis(req: WabisSendRequest) -> JSONResponse:
    """Send a WhatsApp template message via Wabis API."""
    api_key = settings.wabis_api_key
    if not api_key:
        return JSONResponse({"ok": False, "detail": "WABIS_API_KEY not configured on server"}, status_code=503)

    phone_number_id = settings.wabis_phone_number_id
    if not phone_number_id:
        return JSONResponse({"ok": False, "detail": "WABIS_PHONE_NUMBER_ID not configured on server"}, status_code=503)

    phone = req.phone.replace("+", "").replace(" ", "").strip()
    phone_e164 = _normalise_phone(req.phone)
    if not phone:
        return JSONResponse({"ok": False, "detail": "Phone number is required"}, status_code=400)

    # Resolve template_id if not provided
    tid = req.template_id
    if not tid and req.template_name:
        # Look up from current cache
        cached = _cache.get("wabis")
        if cached:
            for t in cached[0]:
                if t["name"] == req.template_name:
                    tid = t["id"]
                    break
        if not tid:
            # Fetch fresh
            from app.wabis_client import _get_template_id  # noqa: PLC0415
            tid = _get_template_id(api_key, req.template_name)

    if not tid:
        return JSONResponse(
            {"ok": False, "detail": f"Template '{req.template_name}' not found in Wabis account"},
            status_code=404,
        )

    send_params: dict[str, Any] = {
        "apiToken": api_key,
        "phone_number_id": phone_number_id,
        "phone_number": phone,
        "template_id": tid,
    }
    if req.params:
        for i, val in enumerate(req.params, 1):
            send_params[f"param{i}"] = str(val)
    if req.header_image_url:
        send_params["header_image_url"] = req.header_image_url

    form_data = urllib.parse.urlencode(send_params).encode()
    send_req = urllib.request.Request(
        f"{WABIS_BASE_URL}/send/template",
        data=form_data,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
            "User-Agent": "PurelevenCRM/1.0",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(send_req, timeout=15, context=_SSL_CONTEXT) as resp:  # noqa: S310
            response_data: dict[str, Any] = json.loads(resp.read())
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        logger.error("Wabis send HTTP %s: %s", exc.code, body[:300])
        _record_status_event(
            message_id=f"wabis_error_{uuid.uuid4().hex[:20]}",
            recipient_phone=phone_e164 or phone,
            status="failed",
            template_name=req.template_name,
            raw_payload={"error": body[:300], "template_id": tid},
            error_code=str(exc.code),
            error_title="Wabis HTTP error",
            error_detail=body[:300],
        )
        return JSONResponse({"ok": False, "detail": body[:300]}, status_code=exc.code)
    except Exception as exc:
        logger.error("Wabis send error: %s", exc)
        _record_status_event(
            message_id=f"wabis_error_{uuid.uuid4().hex[:20]}",
            recipient_phone=phone_e164 or phone,
            status="failed",
            template_name=req.template_name,
            raw_payload={"error": str(exc), "template_id": tid},
            error_title="Wabis send error",
            error_detail=str(exc),
        )
        return JSONResponse({"ok": False, "detail": str(exc)}, status_code=503)

    ok = str(response_data.get("status", "0")) == "1"
    if not ok:
        msg = response_data.get("message", "Unknown Wabis error")
        logger.warning("Wabis logical failure: %s", msg)
        _record_status_event(
            message_id=f"wabis_error_{uuid.uuid4().hex[:20]}",
            recipient_phone=phone_e164 or phone,
            status="failed",
            template_name=req.template_name,
            raw_payload=response_data,
            error_title="Wabis rejected send",
            error_detail=msg,
        )
        return JSONResponse({"ok": False, "detail": msg, "raw": response_data}, status_code=400)

    wamid = response_data.get("wa_message_id", "")
    _record_status_event(
        message_id=wamid or f"wabis_sent_{uuid.uuid4().hex[:20]}",
        recipient_phone=phone_e164 or phone,
        status="sent",
        template_name=req.template_name,
        raw_payload=response_data,
    )
    logger.info("Wabis sent OK: phone=%s template_id=%s wamid=%s", phone, tid, wamid)
    return JSONResponse({"ok": True, "message_id": wamid, "phone": phone, "raw": response_data})


# ─── Send via Meta ────────────────────────────────────────────────────────────

@router.post("/crm/meta-wa/send")
def send_via_meta(req: MetaSendRequest) -> JSONResponse:
    """Send a WhatsApp template message directly via Meta Graph API."""
    token = settings.meta_access_token
    if not token:
        return JSONResponse({"ok": False, "detail": "META_ACCESS_TOKEN not configured on server"}, status_code=503)

    phone_number_id = req.phone_number_id or settings.meta_phone_number_id
    if not phone_number_id:
        return JSONResponse({"ok": False, "detail": "META_PHONE_NUMBER_ID not configured"}, status_code=503)

    phone = req.phone.replace("+", "").replace(" ", "").strip()
    phone_e164 = _normalise_phone(req.phone)
    if not phone:
        return JSONResponse({"ok": False, "detail": "Phone number is required"}, status_code=400)

    template_definition = _meta_template_components(req.template, req.language_code)
    payload: dict[str, Any] = {
        "messaging_product": "whatsapp",
        "to": phone,
        "type": "template",
        "template": {
            "name": req.template,
            "language": {"code": req.language_code},
        },
    }
    components = _build_meta_send_components(template_definition, req.params)
    if components:
        payload["template"]["components"] = components

    url = f"{META_GRAPH_BASE}/{phone_number_id}/messages"
    try:
        resp_data = _http_post(url, payload, headers={"Authorization": f"Bearer {token}"})
    except urllib.error.HTTPError as exc:
        err_body = exc.read().decode("utf-8", errors="replace")
        try:
            detail = json.loads(err_body).get("error", {}).get("message", err_body)
        except Exception:
            detail = err_body[:300]
        _record_status_event(
            message_id=f"meta_error_{uuid.uuid4().hex[:20]}",
            recipient_phone=phone_e164 or phone,
            status="failed",
            template_name=req.template,
            raw_payload={"error": detail, "payload": payload},
            error_code=str(exc.code),
            error_title="Meta HTTP error",
            error_detail=detail,
        )
        return JSONResponse({"ok": False, "detail": detail}, status_code=exc.code)
    except Exception as exc:
        _record_status_event(
            message_id=f"meta_error_{uuid.uuid4().hex[:20]}",
            recipient_phone=phone_e164 or phone,
            status="failed",
            template_name=req.template,
            raw_payload={"error": str(exc), "payload": payload},
            error_title="Meta send error",
            error_detail=str(exc),
        )
        return JSONResponse({"ok": False, "detail": str(exc)}, status_code=503)

    msgs = resp_data.get("messages", [])
    wamid = msgs[0].get("id", "") if msgs else ""
    _record_status_event(
        message_id=wamid or f"meta_sent_{uuid.uuid4().hex[:20]}",
        recipient_phone=phone_e164 or phone,
        status="sent",
        template_name=req.template,
        raw_payload=resp_data,
    )
    logger.info("Meta WA sent OK: phone=%s template=%s wamid=%s", phone, req.template, wamid)
    return JSONResponse({"ok": True, "message_id": wamid, "phone": phone, "raw": resp_data})


# ─── Create Template via Meta WABA ───────────────────────────────────────────

class CreateTemplateRequest(BaseModel):
    name: str
    category: str = "MARKETING"
    language: str = "en_US"
    body: str = ""
    header: str = ""
    footer: str = ""
    buttons: list[dict[str, Any]] = []


@router.post("/crm/whatsapp/create-template")
def create_meta_template(req: CreateTemplateRequest) -> JSONResponse:
    """Submit a new WhatsApp template to Meta WABA for review."""
    token = settings.meta_access_token
    waba_id = getattr(settings, "meta_waba_id", None) or ""

    if not token or not waba_id:
        return JSONResponse(
            {"ok": False, "detail": "META_ACCESS_TOKEN and META_WABA_ID must be set on server"},
            status_code=503,
        )

    components: list[dict[str, Any]] = []
    if req.header.strip():
        header_text = req.header.strip()
        header_component: dict[str, Any] = {"type": "HEADER", "format": "TEXT", "text": header_text}
        header_vars = _count_vars(header_text)
        if header_vars:
            header_component["example"] = {"header_text": _default_meta_examples(header_vars, prefix="Header")}
        components.append(header_component)
    if req.body.strip():
        body_text = req.body.strip()
        body_component: dict[str, Any] = {"type": "BODY", "text": body_text}
        body_vars = _count_vars(body_text)
        if body_vars:
            body_component["example"] = {"body_text": [_default_meta_examples(body_vars)]}
        components.append(body_component)
    if req.footer.strip():
        components.append({"type": "FOOTER", "text": req.footer.strip()})
    if req.buttons:
        buttons: list[dict[str, Any]] = []
        for button in req.buttons:
            next_button = dict(button)
            button_vars = max(_count_vars(str(next_button.get("text", ""))), _count_vars(str(next_button.get("url", ""))))
            if button_vars and "example" not in next_button:
                next_button["example"] = _default_meta_examples(button_vars, prefix="Button")
            buttons.append(next_button)
        components.append({"type": "BUTTONS", "buttons": buttons})

    payload = {
        "name": req.name.strip().lower().replace(" ", "_"),
        "category": req.category.upper(),
        "language": req.language,
        "components": components,
    }

    url = f"{META_GRAPH_BASE}/{waba_id}/message_templates"
    try:
        resp_data = _http_post(url, payload, headers={"Authorization": f"Bearer {token}"})
    except urllib.error.HTTPError as exc:
        err_body = exc.read().decode("utf-8", errors="replace")
        try:
            detail = json.loads(err_body).get("error", {}).get("message", err_body)
        except Exception:
            detail = err_body[:300]
        return JSONResponse({"ok": False, "detail": detail}, status_code=exc.code)
    except Exception as exc:
        return JSONResponse({"ok": False, "detail": str(exc)}, status_code=503)

    # Invalidate meta cache so next sync picks up the new template
    _cache.pop("meta", None)

    return JSONResponse({
        "ok": True,
        "template_id": resp_data.get("id"),
        "name": payload["name"],
        "status": resp_data.get("status", "PENDING"),
        "raw": resp_data,
    })


@router.post("/crm/whatsapp/template/create")
def create_meta_template_legacy(req: CreateTemplateRequest) -> JSONResponse:
    """Backward-compatible alias for older frontend builds."""
    return create_meta_template(req)


# ─── Manual Order Message Trigger ────────────────────────────────────────────

class OrderMessageRequest(BaseModel):
    phone: str = Field(..., description="Customer phone (any format, will be normalised to E.164)")
    order_name: str = Field(..., description="Order name / ID, e.g. PRN-260523-321")
    customer_name: str = Field("", description="Customer first name for template greeting")
    delivery_days: str = Field("3-5 business days", description="Estimated delivery window")
    tracking_id: str = Field("", description="Delhivery AWB number for Track Order button (empty = use order name)")
    shop_domain: str = Field("rwxtic-gz.myshopify.com", description="Shopify shop domain")


@router.post("/crm/whatsapp/trigger-order-message")
def trigger_order_message(req: OrderMessageRequest) -> JSONResponse:
    """
    Manually trigger an order_confirmed_v1 WhatsApp message.
    Use this for orders created outside the Shopify webhook flow (manual / offline orders).
    """
    from app.meta_client import MetaAPIError, send_template_message as meta_send
    from app.wabis_client import WabisError

    # Normalise phone to E.164 digits
    digits = "".join(c for c in req.phone if c.isdigit())
    if len(digits) == 10:
        digits = "91" + digits

    first_name = req.customer_name.strip().split()[0] if req.customer_name.strip() else "there"
    awb = req.tracking_id.strip() or req.order_name.strip()

    # order_confirmed_v1 has 3 body params + 1 URL button param (tracking AWB)
    try:
        result = meta_send(
            phone=digits,
            template_name="order_confirmed_v1",
            language_code="en",
            body_params=[first_name, req.order_name, req.delivery_days],
            button_params=[{
                "sub_type": "url",
                "index": "0",
                "parameters": [{"type": "text", "text": awb}],
            }],
            shop_domain=req.shop_domain,
        )
        success = bool(result.get("messages") or result.get("message_id"))
    except (MetaAPIError, WabisError) as exc:
        logger.error("Order message send failed: phone=%s order=%s error=%s", digits, req.order_name, exc)
        return JSONResponse({"ok": False, "phone": digits, "order": req.order_name, "error": str(exc)}, status_code=500)

    logger.info(
        "Manual order message: phone=%s order=%s success=%s",
        digits, req.order_name, success,
    )
    return JSONResponse({"ok": success, "phone": digits, "order": req.order_name, "result": result})


# ─── Manual Wabis Lead Sync Trigger ──────────────────────────────────────────

@router.post("/crm/whatsapp/sync-wabis-leads", status_code=202)
def trigger_wabis_sync(background_tasks: BackgroundTasks) -> JSONResponse:
    """Manually trigger a Wabis subscriber sync in the background (returns 202 immediately)."""
    from app.services.wabis_lead_sync import sync_wabis_leads
    background_tasks.add_task(sync_wabis_leads)
    return JSONResponse({"ok": True, "message": "Wabis sync started in background"}, status_code=202)
